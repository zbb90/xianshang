#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工时表Excel文件分析器
用于分析工时表的结构、公式和计算逻辑
"""

import pandas as pd
import openpyxl
from openpyxl.formula.translate import Translator
import os
import glob
from pathlib import Path
import json

class ExcelWorkTimeAnalyzer:
    def __init__(self, base_path):
        self.base_path = Path(base_path)
        self.analysis_results = {}
        
    def find_excel_files(self, subfolder="二组"):
        """查找指定子文件夹中的Excel文件"""
        pattern = str(self.base_path / "工时表" / subfolder / "*.xlsx")
        return glob.glob(pattern)
    
    def analyze_single_file(self, file_path):
        """分析单个Excel文件的结构和公式"""
        print(f"\n正在分析文件: {os.path.basename(file_path)}")
        
        try:
            # 使用openpyxl读取工作簿以获取公式信息
            wb = openpyxl.load_workbook(file_path, data_only=False)
            
            # 使用pandas读取数据内容
            excel_data = pd.read_excel(file_path, sheet_name=None, header=None)
            
            file_analysis = {
                'file_name': os.path.basename(file_path),
                'sheet_names': list(excel_data.keys()),
                'worksheets_analysis': {}
            }
            
            # 分析每个工作表
            for sheet_name in wb.sheetnames[:3]:  # 只分析前3个工作表
                print(f"  分析工作表: {sheet_name}")
                
                ws = wb[sheet_name]
                sheet_analysis = self.analyze_worksheet(ws, excel_data.get(sheet_name))
                file_analysis['worksheets_analysis'][sheet_name] = sheet_analysis
            
            return file_analysis
            
        except Exception as e:
            print(f"分析文件 {file_path} 时出错: {str(e)}")
            return None
    
    def analyze_worksheet(self, worksheet, data_df):
        """分析单个工作表的结构和公式"""
        analysis = {
            'dimensions': f"{worksheet.max_row}x{worksheet.max_column}",
            'formulas': [],
            'headers': [],
            'data_types': {},
            'merged_cells': [],
            'sample_data': {}
        }
        
        # 获取合并单元格信息
        for merged_range in worksheet.merged_cells.ranges:
            analysis['merged_cells'].append(str(merged_range))
        
        # 分析前20行的内容
        formulas_found = []
        sample_data = {}
        
        # 简化的方式遍历单元格，避免合并单元格问题
        for row in worksheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=15, values_only=False):
            row_data = {}
            for cell in row:
                try:
                    # 检查cell类型，避免MergedCell
                    if hasattr(cell, 'coordinate') and hasattr(cell, 'value'):
                        cell_ref = cell.coordinate
                        
                        # 记录公式
                        if hasattr(cell, 'data_type') and cell.data_type == 'f' and cell.value:
                            formulas_found.append({
                                'cell': cell_ref,
                                'formula': str(cell.value),
                                'display_value': getattr(cell, 'displayed_value', None)
                            })
                        
                        # 记录样本数据
                        if cell.value is not None:
                            col_letter = cell.column_letter if hasattr(cell, 'column_letter') else f"Col_{cell.column}"
                            row_data[col_letter] = {
                                'value': str(cell.value)[:100],  # 限制长度
                                'data_type': getattr(cell, 'data_type', 'unknown'),
                                'number_format': getattr(cell, 'number_format', 'General')
                            }
                except Exception as e:
                    # 跳过有问题的单元格
                    continue
            
            if row_data:
                row_number = row[0].row if row and hasattr(row[0], 'row') else len(sample_data)+1
                sample_data[f"Row_{row_number}"] = row_data
        
        analysis['formulas'] = formulas_found
        analysis['sample_data'] = sample_data
        
        # 尝试识别表头
        if data_df is not None and not data_df.empty:
            # 查找包含关键字的行作为可能的表头
            keywords = ['日期', '工作内容', '工时', '路程', '地点', '项目', '备注', '时间']
            for idx, row in data_df.iterrows():
                if idx > 10:  # 只检查前10行
                    break
                row_str = ' '.join([str(x) for x in row.values if pd.notna(x)])
                if any(keyword in row_str for keyword in keywords):
                    analysis['headers'].append({
                        'row_index': idx,
                        'content': row.values.tolist()[:10]  # 只取前10列
                    })
        
        return analysis
    
    def analyze_folder(self, subfolder="二组"):
        """分析指定文件夹中的所有Excel文件"""
        excel_files = self.find_excel_files(subfolder)
        print(f"找到 {len(excel_files)} 个Excel文件")
        
        folder_analysis = {
            'folder': subfolder,
            'total_files': len(excel_files),
            'files_analysis': [],
            'common_patterns': {}
        }
        
        # 分析每个文件（限制数量以避免过长）
        for i, file_path in enumerate(excel_files[:5]):  # 只分析前5个文件
            file_analysis = self.analyze_single_file(file_path)
            if file_analysis:
                folder_analysis['files_analysis'].append(file_analysis)
        
        # 分析共同模式
        folder_analysis['common_patterns'] = self.find_common_patterns(folder_analysis['files_analysis'])
        
        return folder_analysis
    
    def find_common_patterns(self, files_analysis):
        """找出文件间的共同模式"""
        patterns = {
            'common_formulas': {},
            'common_headers': {},
            'common_sheet_names': {},
            'formula_types': set()
        }
        
        for file_analysis in files_analysis:
            for sheet_name, sheet_data in file_analysis['worksheets_analysis'].items():
                # 统计工作表名称
                patterns['common_sheet_names'][sheet_name] = patterns['common_sheet_names'].get(sheet_name, 0) + 1
                
                # 分析公式模式
                for formula_info in sheet_data['formulas']:
                    formula = formula_info['formula']
                    
                    # 提取公式类型
                    if formula.startswith('='):
                        func_name = formula.split('(')[0][1:].upper()
                        patterns['formula_types'].add(func_name)
                        patterns['common_formulas'][func_name] = patterns['common_formulas'].get(func_name, 0) + 1
        
        patterns['formula_types'] = list(patterns['formula_types'])
        return patterns
    
    def save_analysis_results(self, analysis_results, output_file):
        """保存分析结果到JSON文件"""
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(analysis_results, f, ensure_ascii=False, indent=2, default=str)
        print(f"分析结果已保存到: {output_file}")

def main():
    # 设置基础路径
    base_path = "/Users/zhaobinbin/Desktop/2025年9月/路径线上化"
    
    # 创建分析器
    analyzer = ExcelWorkTimeAnalyzer(base_path)
    
    # 分析二组文件夹
    print("开始分析二组工时表...")
    analysis_results = analyzer.analyze_folder("二组")
    
    # 保存结果
    output_file = os.path.join(base_path, "工时表分析结果.json")
    analyzer.save_analysis_results(analysis_results, output_file)
    
    # 打印摘要信息
    print("\n=== 分析摘要 ===")
    print(f"总文件数: {analysis_results['total_files']}")
    print(f"分析文件数: {len(analysis_results['files_analysis'])}")
    
    patterns = analysis_results['common_patterns']
    print(f"\n发现的公式类型: {', '.join(patterns['formula_types'])}")
    print(f"常见工作表名称: {list(patterns['common_sheet_names'].keys())}")
    
    if patterns['common_formulas']:
        print("\n最常用的公式函数:")
        for func, count in sorted(patterns['common_formulas'].items(), key=lambda x: x[1], reverse=True)[:5]:
            print(f"  {func}: {count}次")

if __name__ == "__main__":
    main()
