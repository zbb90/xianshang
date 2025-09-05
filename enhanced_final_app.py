#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工时表线上化系统 - 增强最终版
包含门店管理和Excel导出功能
"""

from flask import Flask, request, jsonify, send_file, make_response, session, redirect, url_for
import json
import os
import sqlite3
from datetime import datetime, timedelta
import math
import requests
import time
import hashlib
import urllib.parse
import csv
import io
import zipfile
import secrets
import bcrypt
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
# 设置会话密钥（优先使用环境变量）
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# 导入配置
try:
    from config import AMAP_API_KEY, AMAP_SECRET_KEY
except ImportError:
    # 如果config.py不存在或导入失败，直接从环境变量获取
    AMAP_API_KEY = os.environ.get('AMAP_API_KEY', 'your_amap_api_key_here')
    AMAP_SECRET_KEY = os.environ.get('AMAP_SECRET_KEY', 'your_amap_secret_key_here')

# 高德地图API配置
AMAP_BASE_URL = 'https://restapi.amap.com/v3'

class AmapService:
    """高德地图API服务，具备降级机制"""
    
    @staticmethod
    def _generate_sig(params):
        """生成高德API签名"""
        try:
            # 排序参数
            sorted_params = sorted(params.items())
            # 构建参数字符串
            param_str = '&'.join([f"{k}={v}" for k, v in sorted_params])
            # 添加私钥
            sign_str = param_str + AMAP_SECRET_KEY
            # 生成MD5签名
            sig = hashlib.md5(sign_str.encode('utf-8')).hexdigest()
            return sig
        except Exception as e:
            print(f"🔐 签名生成失败: {e}")
            return None
    
    @staticmethod
    def geocode(address):
        """地理编码：地址转坐标，失败时返回估算坐标"""
        url = f"{AMAP_BASE_URL}/geocode/geo"
        params = {
            'key': AMAP_API_KEY,
            'address': address
        }
        
        # 添加签名
        sig = AmapService._generate_sig(params)
        if sig:
            params['sig'] = sig
        
        try:
            response = requests.get(url, params=params, timeout=5)
            data = response.json()
            
            if data['status'] == '1' and data.get('geocodes'):
                location = data['geocodes'][0]['location'].split(',')
                print(f"✅ 成功获取 {address} 的精确坐标")
                return float(location[0]), float(location[1])  # 经度, 纬度
            else:
                print(f"⚠️ 地理编码API失败: {data.get('info', 'Unknown')}, 使用估算坐标")
                return AmapService._get_fallback_coordinates(address)
        except Exception as e:
            print(f"⚠️ 地理编码请求失败: {e}, 使用估算坐标")
            return AmapService._get_fallback_coordinates(address)
    
    @staticmethod
    def _get_fallback_coordinates(address):
        """根据地址关键词返回估算坐标"""
        # 杭州市区的几个参考点
        fallback_coords = {
            '西湖': (120.1552, 30.2741),
            '萧山': (120.4342, 30.2295), 
            '江干': (120.2102, 30.2901),
            '上城': (120.1692, 30.2444),
            '下城': (120.1619, 30.2756),
            '拱墅': (120.1375, 30.3286),
            '滨江': (120.2119, 30.2084),
            '余杭': (120.3004, 30.4195),
            '火车': (120.2102, 30.2901),  # 火车站
            '机场': (120.4342, 30.2295),  # 机场
        }
        
        for keyword, coords in fallback_coords.items():
            if keyword in address:
                return coords
        
        # 默认返回杭州市中心坐标
        return (120.1552, 30.2741)
    
    @staticmethod
    def calculate_distance_and_time(origin_lng, origin_lat, dest_lng, dest_lat, transport_mode='自驾', route_strategy='fastest'):
        """计算路径距离和时间，具备API降级机制，支持路线策略选择"""
        
        # 首先尝试使用高德地图API
        api_distance, api_time = AmapService._try_amap_api(
            origin_lng, origin_lat, dest_lng, dest_lat, transport_mode, route_strategy
        )
        
        if api_distance is not None:
            print(f"✅ 使用高德地图API计算成功")
            return api_distance, api_time, True
        
        # 禁用备用算法，强制使用高德API
        print(f"❌ 高德地图API不可用，计算失败")
        raise Exception("高德地图API调用失败，请检查API配置或网络连接")
    
    @staticmethod
    def _try_amap_api(origin_lng, origin_lat, dest_lng, dest_lat, transport_mode, route_strategy='fastest'):
        """尝试使用高德地图API - 增强调试版本，支持路线策略选择"""
        try:
            # 根据交通方式选择API
            if transport_mode == '步行':
                url = f"{AMAP_BASE_URL}/direction/walking"
                strategy = 1  # 步行只有一种策略
            else:
                url = f"{AMAP_BASE_URL}/direction/driving"
                # 自驾路线策略：
                # 0: 速度最快 (时间最短)
                # 1: 费用最少 (高速公路费用最少)
                # 2: 距离最短
                # 3: 速度最快，不走高速
                # 4: 躲避拥堵
                # 5: 多策略 (同时使用速度最快和费用最少)
                # 6: 速度最快，且规避收费路段
                # 7: 距离最短，且规避收费路段
                # 8: 躲避拥堵和收费路段
                # 9: 速度最快，躲避拥堵和收费路段
                strategy_map = {
                    'fastest': 0,        # 时间最短 (速度最快)
                    'shortest': 2,       # 距离最短
                    'economical': 1,     # 费用最少
                    'no_highway': 3,     # 速度最快，不走高速
                    'avoid_jam': 4,      # 躲避拥堵
                    'multi': 5,          # 多策略
                    'fastest_free': 6,   # 速度最快，且规避收费路段
                    'shortest_free': 7,  # 距离最短，且规避收费路段
                    'avoid_all': 8,      # 躲避拥堵和收费路段
                    'optimal': 9         # 最优路线 (速度最快，躲避拥堵和收费路段)
                }
                strategy = strategy_map.get(route_strategy, 0)  # 默认使用时间最短
            
            params = {
                'key': AMAP_API_KEY,
                'origin': f"{origin_lng},{origin_lat}",
                'destination': f"{dest_lng},{dest_lat}",
                'strategy': strategy,
                'output': 'json'
            }
            
            # 添加签名
            sig = AmapService._generate_sig(params)
            if sig:
                params['sig'] = sig
            
            print(f"🔄 调用高德API: {transport_mode}")
            print(f"📍 起点: ({origin_lng}, {origin_lat}), 终点: ({dest_lng}, {dest_lat})")
            print(f"🔗 URL: {url}")
            
            response = requests.get(url, params=params, timeout=10)
            
            # 检查HTTP状态码
            if response.status_code != 200:
                print(f"❌ HTTP错误: {response.status_code}")
                return None, None
                
            data = response.json()
            print(f"📦 API响应: status={data.get('status')}, info={data.get('info', 'No info')}")
            
            if data.get('status') == '1' and 'route' in data:
                if 'paths' in data['route'] and data['route']['paths']:
                    path = data['route']['paths'][0]
                    distance_m = int(path['distance'])
                    duration_s = int(path['duration'])
                    
                    distance_km = distance_m / 1000
                    duration_h = duration_s / 3600
                    
                    print(f"✅ 高德API成功: 距离{distance_km:.1f}km, 时间{duration_h:.2f}h")
                    return distance_km, duration_h
            
            # API调用失败的详细信息
            print(f"❌ API调用失败: 状态{data.get('status')}, 信息{data.get('info')}")
            if 'infocode' in data:
                print(f"📝 错误代码: {data['infocode']}")
            
            return None, None
            
        except Exception as e:
            print(f"❌ API请求异常: {e}")
            return None, None
    
    @staticmethod
    def _calculate_fallback(origin_lng, origin_lat, dest_lng, dest_lat, transport_mode):
        """备用距离和时间计算"""
        # 使用球面距离公式计算直线距离
        R = 6371  # 地球半径(km)
        lat1, lon1 = math.radians(origin_lat), math.radians(origin_lng)
        lat2, lon2 = math.radians(dest_lat), math.radians(dest_lng)
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        straight_distance = R * c
        
        # 根据交通方式调整距离系数（实际路径通常比直线距离长）
        distance_factors = {
            '步行': 1.3,    # 步行路径相对较直
            '自驾': 1.4,    # 驾车需要走道路
            '打车': 1.4,    # 同驾车
            '公交': 1.6,    # 公交路线较绕
        }
        
        factor = distance_factors.get(transport_mode, 1.4)
        actual_distance = straight_distance * factor
        
        # 根据交通方式计算时间 - 优化速度设置以更接近高德实际测算
        speed_map = {
            '步行': 4,      # 4 km/h
            '自驾': 60,     # 60 km/h (优化：更接近高德实际测算)
            '打车': 55,     # 55 km/h (考虑等车时间)
            '公交': 25,     # 25 km/h (包含等车和换乘时间)
        }
        
        speed = speed_map.get(transport_mode, 30)
        travel_time = actual_distance / speed
        
        return round(actual_distance, 2), round(travel_time, 2)

def init_db():
    """初始化数据库"""
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            department TEXT NOT NULL,
            position TEXT NOT NULL,
            group_name TEXT,
            username TEXT UNIQUE,
            email TEXT UNIQUE,
            password_hash TEXT,
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 为现有用户添加默认认证信息（如果列不存在）
    cursor.execute("PRAGMA table_info(users)")
    columns = [column[1] for column in cursor.fetchall()]
    
    # 添加group_name列（如果不存在）
    if 'group_name' not in columns:
        cursor.execute('ALTER TABLE users ADD COLUMN group_name TEXT')
    
    if 'username' not in columns:
        cursor.execute('ALTER TABLE users ADD COLUMN username TEXT')
    if 'email' not in columns:
        cursor.execute('ALTER TABLE users ADD COLUMN email TEXT')
    if 'password_hash' not in columns:
        cursor.execute('ALTER TABLE users ADD COLUMN password_hash TEXT')
    if 'is_active' not in columns:
        cursor.execute('ALTER TABLE users ADD COLUMN is_active BOOLEAN DEFAULT 1')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS locations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            address TEXT NOT NULL,
            longitude REAL,
            latitude REAL,
            geocoded BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 新增门店表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS stores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            store_code TEXT NOT NULL UNIQUE,
            store_name TEXT NOT NULL,
            store_city TEXT NOT NULL,
            longitude REAL NOT NULL,
            latitude REAL NOT NULL,
            address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS timesheet_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date DATE NOT NULL,
            location_from_id INTEGER,
            location_to_id INTEGER,
            store_from_id INTEGER,
            store_to_id INTEGER,
            transport_mode TEXT DEFAULT '自驾',
            distance REAL,
            travel_time REAL,
            work_hours REAL NOT NULL,
            notes TEXT,
            api_used BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            -- 新增字段支持简化工时录入
            store_code TEXT,
            store_name TEXT,
            work_date DATE,
            start_time TEXT,
            end_time TEXT,
            work_content TEXT,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (location_from_id) REFERENCES locations (id),
            FOREIGN KEY (location_to_id) REFERENCES locations (id),
            FOREIGN KEY (store_from_id) REFERENCES stores (id),
            FOREIGN KEY (store_to_id) REFERENCES stores (id)
        )
    ''')
    
    # 检查并添加新字段
    cursor.execute("PRAGMA table_info(timesheet_records)")
    columns = [col[1] for col in cursor.fetchall()]
    
    new_columns = [
        ('store_code', 'TEXT'),
        ('store_name', 'TEXT'),
        ('work_date', 'DATE'),
        ('start_time', 'TEXT'),
        ('end_time', 'TEXT'),
        ('work_content', 'TEXT')
    ]
    
    for col_name, col_type in new_columns:
        if col_name not in columns:
            cursor.execute(f'ALTER TABLE timesheet_records ADD COLUMN {col_name} {col_type}')
    
    # 插入示例数据
    cursor.execute("SELECT COUNT(*) FROM users")
    if cursor.fetchone()[0] == 0:
        sample_users = [
            ('张三', '稽核四组', '稽核专员'),
            ('李四', '稽核四组', '稽核主管'),
            ('王五', '稽核二组', '稽核专员'),
            ('赵六', '稽核二组', '稽核专员'),
        ]
        cursor.executemany("INSERT INTO users (name, department, position) VALUES (?, ?, ?)", sample_users)
        
        # 插入常用地点
        sample_locations = [
            ('公司总部', '杭州市西湖区文三路288号', None, None, False),
            ('杭州火车东站', '杭州市江干区天城路1号杭州东站', None, None, False),
            ('萧山国际机场', '杭州市萧山区萧山国际机场', None, None, False),
            ('市民中心', '杭州市江干区解放东路18号', None, None, False),
            ('西湖景区', '杭州市西湖区西湖', None, None, False),
            ('滨江办事处', '杭州市滨江区江南大道', None, None, False),
        ]
        cursor.executemany("INSERT INTO locations (name, address, longitude, latitude, geocoded) VALUES (?, ?, ?, ?, ?)", sample_locations)
        
        # 插入示例门店
        sample_stores = [
            ('HZ001', '杭州西湖店', '杭州市', 120.1552, 30.2741, '杭州市西湖区文三路100号'),
            ('HZ002', '杭州萧山店', '杭州市', 120.4342, 30.2295, '杭州市萧山区市心路50号'),
            ('HZ003', '杭州滨江店', '杭州市', 120.2119, 30.2084, '杭州市滨江区江南大道200号'),
            ('SH001', '上海徐汇店', '上海市', 121.4737, 31.2304, '上海市徐汇区漕溪北路88号'),
            ('BJ001', '北京朝阳店', '北京市', 116.4074, 39.9042, '北京市朝阳区建国路120号'),
        ]
        cursor.executemany("INSERT INTO stores (store_code, store_name, store_city, longitude, latitude, address) VALUES (?, ?, ?, ?, ?, ?)", sample_stores)
    
    conn.commit()
    conn.close()
    print("✅ 数据库初始化完成")

@app.route('/login')
def login_page():
    return '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>用户登录 - 智能工时表管理系统</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        .login-container {
            max-width: 400px;
            margin: 100px auto;
            padding: 2rem;
            box-shadow: 0 0 20px rgba(0,0,0,0.1);
            border-radius: 10px;
        }
        .system-title {
            text-align: center;
            color: #2c3e50;
            margin-bottom: 2rem;
        }
    </style>
</head>
<body class="bg-light">
    <div class="container">
        <div class="login-container bg-white">
            <h2 class="system-title">🚀 智能工时表管理系统</h2>
            <h4 class="text-center mb-4">用户登录</h4>
            
            <div id="message" class="alert" style="display: none;"></div>
            
            <form id="loginForm">
                <div class="mb-3">
                    <label for="name" class="form-label">姓名</label>
                    <input type="text" class="form-control" id="name" required placeholder="请输入您的姓名">
                </div>
                <div class="mb-3">
                    <label for="group" class="form-label">组别</label>
                    <select class="form-control" id="group" required>
                        <option value="">请选择组别</option>
                        <option value="稽核一组">稽核一组</option>
                        <option value="稽核二组">稽核二组</option>
                        <option value="稽核三组">稽核三组</option>
                        <option value="稽核四组">稽核四组</option>
                    </select>
                </div>
                <div class="mb-3">
                    <label for="password" class="form-label">密码</label>
                    <input type="password" class="form-control" id="password" required>
                </div>
                <button type="submit" class="btn btn-primary w-100 mb-3">登录</button>
            </form>
            
            <div class="text-center">
                <p>还没有账户？ <a href="/register">立即注册</a></p>
            </div>
        </div>
    </div>

    <script>
        document.getElementById('loginForm').addEventListener('submit', async (e) => {
            e.preventDefault();
            
            const name = document.getElementById('name').value;
            const group = document.getElementById('group').value;
            const password = document.getElementById('password').value;
            const messageDiv = document.getElementById('message');
            
            if (!name || !group || !password) {
                messageDiv.className = 'alert alert-warning';
                messageDiv.textContent = '请填写完整信息';
                messageDiv.style.display = 'block';
                return;
            }
            
            try {
                const response = await fetch('/api/login', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({ 
                        name: name,
                        group: group,
                        password: password 
                    })
                });
                
                const result = await response.json();
                
                if (result.success) {
                    messageDiv.className = 'alert alert-success';
                    messageDiv.textContent = '登录成功，正在跳转...';
                    messageDiv.style.display = 'block';
                    
                    setTimeout(() => {
                        window.location.href = '/dashboard';
                    }, 1000);
                } else {
                    messageDiv.className = 'alert alert-danger';
                    messageDiv.textContent = result.error || '登录失败';
                    messageDiv.style.display = 'block';
                }
            } catch (error) {
                messageDiv.className = 'alert alert-danger';
                messageDiv.textContent = '网络错误，请重试';
                messageDiv.style.display = 'block';
            }
        });
    </script>
</body>
</html>
    '''

@app.route('/version')
def version_info():
    return jsonify({
        'version': 'v2.0',
        'description': '简化注册系统 - 姓名+组别+密码',
        'timestamp': datetime.now().isoformat()
    })

@app.route('/register')
def register_page():
    return '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>用户注册 - 智能工时表管理系统</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
    <meta http-equiv="Pragma" content="no-cache">
    <meta http-equiv="Expires" content="0">
    <style>
        .register-container {
            max-width: 500px;
            margin: 50px auto;
            padding: 2rem;
            box-shadow: 0 0 20px rgba(0,0,0,0.1);
            border-radius: 10px;
        }
        .system-title {
            text-align: center;
            color: #2c3e50;
            margin-bottom: 2rem;
        }
    </style>
</head>
<body class="bg-light">
    <div class="container">
        <div class="register-container bg-white">
            <h2 class="system-title">🚀 智能工时表管理系统</h2>
            <h4 class="text-center mb-4">用户注册</h4>
            <p class="text-center text-muted small">版本: v2.0 - 简化注册</p>
            
            <div id="message" class="alert" style="display: none;"></div>
            
            <form id="registerForm">
                <div class="mb-3">
                    <label for="name" class="form-label">姓名</label>
                    <input type="text" class="form-control" id="name" required placeholder="请输入您的真实姓名">
                </div>
                <div class="mb-3">
                    <label for="group" class="form-label">组别</label>
                    <select class="form-control" id="group" required>
                        <option value="">请选择组别</option>
                        <option value="稽核一组">稽核一组</option>
                        <option value="稽核二组">稽核二组</option>
                        <option value="稽核三组">稽核三组</option>
                        <option value="稽核四组">稽核四组</option>
                    </select>
                </div>
                <div class="mb-3">
                    <label for="password" class="form-label">密码</label>
                    <input type="password" class="form-control" id="password" required minlength="6" placeholder="请设置登录密码（至少6位）">
                </div>
                <div class="mb-3">
                    <label for="role" class="form-label">角色</label>
                    <select class="form-control" id="role" required>
                        <option value="">请选择角色</option>
                        <option value="普通用户">普通用户（只能录入工时）</option>
                        <option value="管理员">管理员（可管理门店信息）</option>
                    </select>
                </div>
                <button type="submit" class="btn btn-success w-100 mb-3">注册</button>
            </form>
            
            <div class="text-center">
                <p>已有账户？ <a href="/login">立即登录</a></p>
            </div>
        </div>
    </div>

    <script>
        document.getElementById('registerForm').addEventListener('submit', async (e) => {
            e.preventDefault();
            
            const formData = {
                name: document.getElementById('name').value,
                group: document.getElementById('group').value,
                password: document.getElementById('password').value,
                role: document.getElementById('role').value
            };
            
            const messageDiv = document.getElementById('message');
            
            try {
                const response = await fetch('/api/register', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify(formData)
                });
                
                const result = await response.json();
                
                if (result.success) {
                    messageDiv.className = 'alert alert-success';
                    messageDiv.textContent = '注册成功！正在跳转到登录页面...';
                    messageDiv.style.display = 'block';
                    
                    setTimeout(() => {
                        window.location.href = '/login';
                    }, 2000);
                } else {
                    messageDiv.className = 'alert alert-danger';
                    messageDiv.textContent = result.error || '注册失败';
                    messageDiv.style.display = 'block';
                }
            } catch (error) {
                messageDiv.className = 'alert alert-danger';
                messageDiv.textContent = '网络错误，请重试';
                messageDiv.style.display = 'block';
            }
        });
    </script>
</body>
</html>
    '''

# 认证装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': '请先登录', 'redirect': '/login'}), 401
        
        conn = sqlite3.connect('enhanced_timesheet.db')
        cursor = conn.cursor()
        cursor.execute("SELECT position FROM users WHERE id = ?", (session['user_id'],))
        user = cursor.fetchone()
        conn.close()
        
        if not user or user[0] not in ['管理员', '系统管理员']:
            return jsonify({'error': '权限不足'}), 403
        return f(*args, **kwargs)
    return decorated_function

@app.route('/dashboard')
@login_required
def dashboard():
    """管理控制台主页"""
    return '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>智能工时表管理系统</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        body { 
            background-color: #f8f9fa; 
            margin: 0;
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        }
        .sidebar {
            width: 250px;
            background: linear-gradient(180deg, #2c3e50 0%, #34495e 100%);
            min-height: 100vh;
            position: fixed;
            left: 0;
            top: 0;
            z-index: 1000;
            color: white;
            overflow-y: auto;
        }
        .sidebar-header {
            padding: 20px;
            text-align: center;
            border-bottom: 1px solid #34495e;
        }
        .sidebar-header h4 {
            color: white;
            margin: 0;
            font-size: 18px;
        }
        .nav-menu {
            padding: 20px 0;
        }
        .nav-item {
            margin: 5px 0;
        }
        .nav-link {
            color: #bdc3c7 !important;
            padding: 12px 20px;
            display: flex;
            align-items: center;
            text-decoration: none;
            transition: all 0.3s;
            border: none;
            background: none;
        }
        .nav-link:hover, .nav-link.active {
            background-color: #3498db;
            color: white !important;
        }
        .nav-link i {
            width: 20px;
            margin-right: 10px;
        }
        .main-content {
            margin-left: 250px;
            padding: 0;
            min-height: 100vh;
        }
        .top-bar {
            background: white;
            padding: 15px 30px;
            border-bottom: 1px solid #e9ecef;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .content-area {
            padding: 30px;
        }
        .page-title {
            margin: 0 0 20px 0;
            color: #2c3e50;
            font-size: 24px;
            font-weight: 600;
        }
        .user-info {
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .user-avatar {
            width: 32px;
            height: 32px;
            background: #3498db;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: bold;
        }
        .card {
            border: none;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        }
        .card-header {
            background: #f8f9fa;
            border-bottom: 1px solid #e9ecef;
            font-weight: 600;
        }
        .hidden { display: none; }
    </style>
</head>
<body>
    <div class="sidebar">
        <div class="sidebar-header">
            <h4><i class="fas fa-clock"></i> 工时管理系统</h4>
        </div>
        <nav class="nav-menu">
            <div class="nav-item">
                <button class="nav-link active w-100" onclick="showPage('timesheet')">
                    <i class="fas fa-clock"></i> 工时录入
                </button>
            </div>
            <div class="nav-item admin-only">
                <button class="nav-link w-100" onclick="showPage('stores')">
                    <i class="fas fa-store"></i> 门店管理
                </button>
            </div>
            <div class="nav-item">
                <button class="nav-link w-100" onclick="showPage('reports')">
                    <i class="fas fa-chart-bar"></i> 数据报表
                </button>
            </div>
            <div class="nav-item admin-only">
                <button class="nav-link w-100" onclick="showPage('users')">
                    <i class="fas fa-users"></i> 用户管理
                </button>
            </div>
            <div class="nav-item">
                <button class="nav-link w-100" onclick="logout()">
                    <i class="fas fa-sign-out-alt"></i> 退出登录
                </button>
            </div>
        </nav>
    </div>

    <div class="main-content">
        <div class="top-bar">
            <h1 class="page-title" id="pageTitle">工时录入</h1>
            <div class="user-info">
                <div class="user-avatar" id="userAvatar"></div>
                <div>
                    <div id="userName" style="font-weight: 600;"></div>
                    <div id="userRole" style="font-size: 12px; color: #6c757d;"></div>
                </div>
            </div>
        </div>

        <div class="content-area">
            <!-- 工时录入页面 -->
            <div id="timesheet-page" class="page-content">
                <div class="card">
                    <div class="card-header">
                        <h5 class="mb-0">工时录入</h5>
                    </div>
                    <div class="card-body">
                        <form id="timesheetForm">
                            <div class="row">
                                <div class="col-md-6 mb-3">
                                    <label class="form-label">门店编码</label>
                                    <select class="form-control" id="storeCode" required>
                                        <option value="">请选择门店</option>
                                    </select>
                                </div>
                                <div class="col-md-6 mb-3">
                                    <label class="form-label">工作日期</label>
                                    <input type="date" class="form-control" id="workDate" required>
                                </div>
                            </div>
                            <div class="row">
                                <div class="col-md-6 mb-3">
                                    <label class="form-label">开始时间</label>
                                    <input type="time" class="form-control" id="startTime" required>
                                </div>
                                <div class="col-md-6 mb-3">
                                    <label class="form-label">结束时间</label>
                                    <input type="time" class="form-control" id="endTime" required>
                                </div>
                            </div>
                            <div class="mb-3">
                                <label class="form-label">工作内容</label>
                                <textarea class="form-control" id="workContent" rows="3" placeholder="请简要描述工作内容"></textarea>
                            </div>
                            <button type="submit" class="btn btn-primary">
                                <i class="fas fa-save"></i> 保存工时记录
                            </button>
                        </form>
                    </div>
                </div>
            </div>

            <!-- 门店管理页面 -->
            <div id="stores-page" class="page-content hidden">
                <div class="card">
                    <div class="card-header d-flex justify-content-between align-items-center">
                        <h5 class="mb-0">门店管理</h5>
                        <button class="btn btn-success" onclick="showImportModal()">
                            <i class="fas fa-upload"></i> 导入门店信息
                        </button>
                    </div>
                    <div class="card-body">
                        <div class="table-responsive">
                            <table class="table table-striped">
                                <thead>
                                    <tr>
                                        <th>门店编码</th>
                                        <th>门店名称</th>
                                        <th>城市</th>
                                        <th>地址</th>
                                        <th>状态</th>
                                        <th>操作</th>
                                    </tr>
                                </thead>
                                <tbody id="storesTableBody">
                                    <!-- 门店数据将在这里显示 -->
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>

            <!-- 其他页面内容 -->
            <div id="reports-page" class="page-content hidden">
                <div class="card">
                    <div class="card-header">
                        <h5 class="mb-0">数据报表</h5>
                    </div>
                    <div class="card-body">
                        <p>数据报表功能开发中...</p>
                    </div>
                </div>
            </div>

            <div id="users-page" class="page-content hidden">
                <div class="card">
                    <div class="card-header">
                        <h5 class="mb-0">用户管理</h5>
                    </div>
                    <div class="card-body">
                        <p>用户管理功能开发中...</p>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- 门店导入模态框 -->
    <div class="modal fade" id="importModal" tabindex="-1">
        <div class="modal-dialog">
            <div class="modal-content">
                <div class="modal-header">
                    <h5 class="modal-title">导入门店信息</h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body">
                    <form id="importForm">
                        <div class="mb-3">
                            <label class="form-label">选择Excel文件</label>
                            <input type="file" class="form-control" id="storeFile" accept=".xlsx,.xls" required>
                        </div>
                        <div class="alert alert-info">
                            <small>请确保Excel文件包含：门店编码、门店名称、城市、地址等列</small>
                        </div>
                    </form>
                </div>
                <div class="modal-footer">
                    <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">取消</button>
                    <button type="button" class="btn btn-primary" onclick="importStores()">导入</button>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        let currentUser = null;

        // 页面加载时获取用户信息
        document.addEventListener('DOMContentLoaded', async () => {
            await loadUserInfo();
            await loadStores();
            setDefaultDate();
        });

        async function loadUserInfo() {
            try {
                const response = await fetch('/api/user/info');
                const user = await response.json();
                currentUser = user;
                
                document.getElementById('userName').textContent = user.name;
                document.getElementById('userRole').textContent = `${user.group_name} - ${user.position}`;
                document.getElementById('userAvatar').textContent = user.name.charAt(0);

                // 根据权限显示/隐藏管理员功能
                if (user.position !== '管理员') {
                    document.querySelectorAll('.admin-only').forEach(el => {
                        el.style.display = 'none';
                    });
                }
            } catch (error) {
                console.error('获取用户信息失败:', error);
                window.location.href = '/login';
            }
        }

        async function loadStores() {
            try {
                const response = await fetch('/api/stores');
                const stores = await response.json();
                
                const storeSelect = document.getElementById('storeCode');
                storeSelect.innerHTML = '<option value="">请选择门店</option>';
                
                stores.forEach(store => {
                    const option = document.createElement('option');
                    option.value = store.store_code;
                    option.textContent = `${store.store_code} - ${store.store_name}`;
                    storeSelect.appendChild(option);
                });

                // 更新门店管理表格
                updateStoresTable(stores);
            } catch (error) {
                console.error('加载门店信息失败:', error);
            }
        }

        function updateStoresTable(stores) {
            const tbody = document.getElementById('storesTableBody');
            tbody.innerHTML = '';
            
            stores.forEach(store => {
                const row = document.createElement('tr');
                row.innerHTML = `
                    <td>${store.store_code}</td>
                    <td>${store.store_name}</td>
                    <td>${store.city}</td>
                    <td>${store.address || '-'}</td>
                    <td><span class="badge bg-success">正常</span></td>
                    <td>
                        <button class="btn btn-sm btn-outline-primary">编辑</button>
                        <button class="btn btn-sm btn-outline-danger">删除</button>
                    </td>
                `;
                tbody.appendChild(row);
            });
        }

        function setDefaultDate() {
            const today = new Date().toISOString().split('T')[0];
            document.getElementById('workDate').value = today;
        }

        function showPage(pageId) {
            // 隐藏所有页面
            document.querySelectorAll('.page-content').forEach(page => {
                page.classList.add('hidden');
            });
            
            // 移除所有导航链接的active状态
            document.querySelectorAll('.nav-link').forEach(link => {
                link.classList.remove('active');
            });
            
            // 显示选中的页面
            document.getElementById(pageId + '-page').classList.remove('hidden');
            
            // 设置active状态
            event.target.classList.add('active');
            
            // 更新页面标题
            const titles = {
                'timesheet': '工时录入',
                'stores': '门店管理',
                'reports': '数据报表',
                'users': '用户管理'
            };
            document.getElementById('pageTitle').textContent = titles[pageId];
        }

        function showImportModal() {
            new bootstrap.Modal(document.getElementById('importModal')).show();
        }

        async function importStores() {
            const fileInput = document.getElementById('storeFile');
            const file = fileInput.files[0];
            
            if (!file) {
                alert('请选择文件');
                return;
            }

            const formData = new FormData();
            formData.append('file', file);

            try {
                const response = await fetch('/api/stores/import', {
                    method: 'POST',
                    body: formData
                });

                const result = await response.json();
                if (result.success) {
                    alert('门店信息导入成功');
                    bootstrap.Modal.getInstance(document.getElementById('importModal')).hide();
                    await loadStores();
                } else {
                    alert('导入失败: ' + result.error);
                }
            } catch (error) {
                alert('导入失败: ' + error.message);
            }
        }

        // 工时表单提交
        document.getElementById('timesheetForm').addEventListener('submit', async (e) => {
            e.preventDefault();
            
            const formData = {
                store_code: document.getElementById('storeCode').value,
                work_date: document.getElementById('workDate').value,
                start_time: document.getElementById('startTime').value,
                end_time: document.getElementById('endTime').value,
                work_content: document.getElementById('workContent').value
            };
            
            try {
                const response = await fetch('/api/timesheet', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify(formData)
                });
                
                const result = await response.json();
                if (result.success) {
                    alert('工时记录保存成功！');
                    document.getElementById('timesheetForm').reset();
                    setDefaultDate();
                } else {
                    alert('保存失败: ' + result.error);
                }
            } catch (error) {
                alert('网络错误: ' + error.message);
            }
        });

        async function logout() {
            try {
                await fetch('/api/logout', { method: 'POST' });
                window.location.href = '/login';
            } catch (error) {
                window.location.href = '/login';
            }
        }
    </script>
</body>
</html>
    '''

@app.route('/')
def index():
    return '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>智能工时表管理系统</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'PingFang SC', 'Helvetica Neue', Arial, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
            min-height: 100vh; 
            padding: 20px; 
        }
        .container { max-width: 1200px; margin: 0 auto; }
        .header { 
            background: rgba(255,255,255,0.95); 
            color: #2c3e50; 
            padding: 30px; 
            text-align: center; 
            border-radius: 16px; 
            margin-bottom: 30px; 
            backdrop-filter: blur(10px);
            box-shadow: 0 8px 32px rgba(0,0,0,0.1);
        }
        .header h1 { font-size: 2.5em; font-weight: 300; margin-bottom: 10px; }
        .header .subtitle { font-size: 1.1em; opacity: 0.7; margin-bottom: 15px; }
        .nav-tabs {
            display: flex;
            background: rgba(255,255,255,0.9);
            border-radius: 12px;
            padding: 8px;
            margin: 0 auto;
            max-width: 600px;
        }
        .nav-tab {
            flex: 1;
            padding: 12px 20px;
            text-align: center;
            border-radius: 8px;
            cursor: pointer;
            transition: all 0.3s ease;
            font-weight: 500;
        }
        .nav-tab.active {
            background: linear-gradient(135deg, #667eea, #764ba2);
            color: white;
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
        }
        .nav-tab:not(.active):hover {
            background: rgba(102, 126, 234, 0.1);
        }
        .card { 
            background: rgba(255,255,255,0.95); 
            padding: 30px; 
            border-radius: 16px; 
            margin-bottom: 25px; 
            backdrop-filter: blur(10px);
            box-shadow: 0 8px 32px rgba(0,0,0,0.1);
            display: none;
        }
        .card.active { display: block; }
        .card h3 { color: #2c3e50; margin-bottom: 20px; font-size: 1.3em; }
        .form-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; }
        .form-group { margin-bottom: 20px; }
        .form-group label { 
            display: block; 
            margin-bottom: 8px; 
            font-weight: 600; 
            color: #2c3e50; 
            font-size: 0.95em;
        }
        .form-control { 
            width: 100%; 
            padding: 12px 16px; 
            border: 2px solid #e1e8ed; 
            border-radius: 8px; 
            font-size: 16px; 
            transition: all 0.3s ease;
            background: white;
        }
        .form-control:focus { 
            border-color: #667eea; 
            outline: none; 
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1); 
        }
        .btn { 
            padding: 12px 24px; 
            border: none; 
            border-radius: 8px; 
            cursor: pointer; 
            font-size: 16px; 
            font-weight: 600; 
            transition: all 0.3s ease;
            margin-right: 10px;
            margin-bottom: 10px;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }
        .btn-primary { background: linear-gradient(135deg, #667eea, #764ba2); color: white; }
        .btn-success { background: linear-gradient(135deg, #00b894, #00cec9); color: white; }
        .btn-info { background: linear-gradient(135deg, #74b9ff, #0984e3); color: white; }
        .btn-warning { background: linear-gradient(135deg, #fdcb6e, #e17055); color: white; }
        .btn-secondary { background: #6c757d; color: white; }
        .btn:hover { transform: translateY(-2px); box-shadow: 0 8px 25px rgba(0,0,0,0.15); }
        .btn:disabled { opacity: 0.6; cursor: not-allowed; transform: none; }
        .route-result { 
            background: linear-gradient(135deg, #e8f5e8, #f0fff0); 
            border: 2px solid #00b894; 
            padding: 20px; 
            border-radius: 12px; 
            margin-top: 20px;
        }
        .route-header { 
            font-weight: 600; 
            color: #00b894; 
            margin-bottom: 15px; 
            font-size: 1.1em;
        }
        .route-details { 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); 
            gap: 15px; 
        }
        .route-item { text-align: center; }
        .route-value { font-size: 1.3em; font-weight: bold; color: #00b894; }
        .route-label { font-size: 0.9em; opacity: 0.8; margin-top: 5px; }
        .message { 
            padding: 15px 20px; 
            border-radius: 8px; 
            margin-bottom: 20px; 
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .message.success { background: #d4edda; border: 2px solid #c3e6cb; color: #155724; }
        .message.error { background: #f8d7da; border: 2px solid #f5c6cb; color: #721c24; }
        .message.info { background: #d1ecf1; border: 2px solid #bee5eb; color: #0c5460; }
        
        .store-name-display {
            margin-top: 8px;
            padding: 8px 12px;
            background-color: #f8f9fa;
            border: 1px solid #e9ecef;
            border-radius: 6px;
            font-size: 14px;
            color: #495057;
            min-height: 20px;
            transition: all 0.3s ease;
        }
        
        .store-name-display.found {
            background-color: #d4edda;
            border-color: #c3e6cb;
            color: #155724;
        }
        
        .store-name-display.not-found {
            background-color: #f8d7da;
            border-color: #f5c6cb;
            color: #721c24;
        }
        
        .store-name-display.empty {
            display: none;
        }
        
        /* 加载动画 */
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        /* 路线策略选择框优化 */
        #routeStrategyGroup {
            transition: all 0.3s ease;
        }
        
        #routeStrategy option {
            padding: 8px 12px;
            font-size: 14px;
        }
        
        .stats-grid { 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
            gap: 20px; 
            margin-bottom: 30px; 
        }
        .stat-card { 
            background: linear-gradient(135deg, #667eea, #764ba2); 
            color: white; 
            padding: 25px; 
            border-radius: 12px; 
            text-align: center;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }
        .stat-number { font-size: 2.2em; font-weight: bold; margin-bottom: 8px; }
        .stat-label { opacity: 0.9; font-size: 0.95em; }
        .table-container { overflow-x: auto; border-radius: 12px; }
        .table { 
            width: 100%; 
            border-collapse: collapse; 
            background: white; 
            border-radius: 12px; 
            overflow: hidden;
        }
        .table th, .table td { padding: 15px; text-align: left; border-bottom: 1px solid #e1e8ed; }
        .table th { background: #f8f9fa; font-weight: 600; color: #2c3e50; }
        .table tr:hover { background: #f8f9ff; }
        .loading { 
            display: inline-block; 
            width: 18px; 
            height: 18px; 
            border: 2px solid #f3f3f3; 
            border-top: 2px solid #667eea; 
            border-radius: 50%; 
            animation: spin 1s linear infinite; 
        }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
        .badge { 
            display: inline-block; 
            padding: 4px 8px; 
            border-radius: 12px; 
            font-size: 0.8em; 
            font-weight: 500;
        }
        .badge-api { background: #e8f5e8; color: #00b894; }
        .badge-fallback { background: #fff3cd; color: #856404; }
        .file-upload {
            border: 2px dashed #667eea;
            border-radius: 8px;
            padding: 30px;
            text-align: center;
            background: #f8f9ff;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        .file-upload:hover {
            border-color: #764ba2;
            background: #f0f4ff;
        }
        .file-upload.dragover {
            border-color: #00b894;
            background: #e8f5e8;
        }
        @media (max-width: 768px) { 
            .form-grid { grid-template-columns: 1fr; }
            .stats-grid { grid-template-columns: repeat(2, 1fr); }
            .route-details { grid-template-columns: repeat(2, 1fr); }
            .nav-tabs { flex-direction: column; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="d-flex justify-content-between align-items-center">
                <div>
                    <h1>🚀 智能工时表管理系统</h1>
                    <div class="subtitle">门店管理 + 工时记录 + 数据导出</div>
                </div>
                <div class="user-info">
                    <span class="text-muted me-3">当前用户：<strong id="currentUser">加载中...</strong></span>
                    <button type="button" class="btn btn-outline-danger btn-sm" onclick="logout()">退出登录</button>
                </div>
            </div>
            <div class="nav-tabs">
                <div class="nav-tab active" onclick="switchTab('timesheet')">⏰ 工时管理</div>
                <div class="nav-tab" onclick="switchTab('stores')">🏪 门店管理</div>
                <div class="nav-tab" onclick="switchTab('export')">📊 数据导出</div>
            </div>
        </div>

        <div id="message"></div>

        <!-- 工时管理面板 -->
        <div id="timesheet-panel" class="card active">
            <!-- 用户选择 -->
            <div id="userPanel">
                <h3>👥 用户选择</h3>
                <div class="form-grid">
                    <div class="form-group">
                        <label>选择用户</label>
                        <select id="userSelect" class="form-control">
                            <option value="">请选择用户...</option>
                        </select>
                    </div>
                </div>
                <button onclick="selectUser()" class="btn btn-primary">
                    <span>✅</span>开始工作
                </button>
                <button onclick="showUserStats()" class="btn btn-info">
                    <span>📊</span>查看统计
                </button>
            </div>

            <!-- 统计面板 -->
            <div class="stats-grid" id="statsPanel" style="display: none; margin-top: 20px;">
            </div>

            <!-- 主工作面板 -->
            <div id="workPanel" style="display: none;">
                <h3 id="workTitle">📝 工时录入</h3>
                
                <div class="form-grid">
                    <div class="form-group">
                        <label>📅 工作日期</label>
                        <input type="date" id="workDate" class="form-control">
                    </div>
                    <div class="form-group">
                        <label>⏰ 工作时长 (小时)</label>
                        <input type="number" id="workHours" class="form-control" value="8" step="0.5" min="0" max="24">
                    </div>
                </div>

                <div class="form-grid">
                    <div class="form-group">
                        <label>📍 出发地点类型</label>
                        <select id="fromType" class="form-control" onchange="toggleLocationOptions('from')">
                            <option value="location">🏢 常用地点</option>
                            <option value="store">🏪 门店</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>🎯 目的地点类型</label>
                        <select id="toType" class="form-control" onchange="toggleLocationOptions('to')">
                            <option value="location">🏢 常用地点</option>
                            <option value="store">🏪 门店</option>
                        </select>
                    </div>
                </div>

                <div class="form-grid">
                    <div class="form-group">
                        <label id="fromLabel">📍 出发地点</label>
                        <select id="fromLocation" class="form-control" onchange="autoCalculate()">
                            <option value="">选择出发地点...</option>
                        </select>
                        <div style="display: none;" id="fromStoreContainer">
                            <input type="text" id="fromStoreCode" class="form-control" placeholder="请输入出发门店编码..." maxlength="10" oninput="onStoreCodeInput('from')">
                            <div id="fromStoreName" class="store-name-display"></div>
                        </div>
                    </div>
                    <div class="form-group">
                        <label id="toLabel">🎯 目的地点</label>
                        <select id="toLocation" class="form-control" onchange="autoCalculate()">
                            <option value="">选择目的地点...</option>
                        </select>
                        <div style="display: none;" id="toStoreContainer">
                            <input type="text" id="toStoreCode" class="form-control" placeholder="请输入目的门店编码..." maxlength="10" oninput="onStoreCodeInput('to')">
                            <div id="toStoreName" class="store-name-display"></div>
                        </div>
                    </div>
                </div>

                <div class="form-grid">
                    <div class="form-group">
                        <label>🚗 交通方式</label>
                        <select id="transportMode" class="form-control" onchange="onTransportModeChange()">
                            <option value="自驾">🚗 自驾</option>
                            <option value="步行">🚶 步行</option>
                            <option value="打车">🚕 打车</option>
                            <option value="公交">🚌 公交</option>
                        </select>
                    </div>
                    <div class="form-group" id="routeStrategyGroup">
                        <label>🎯 路线策略</label>
                        <select id="routeStrategy" class="form-control" onchange="autoCalculate()">
                            <option value="fastest">⚡ 时间最短 (速度最快)</option>
                            <option value="optimal">🎯 最优路线 (躲避拥堵)</option>
                            <option value="shortest">📏 距离最短</option>
                            <option value="economical">💰 费用最少</option>
                            <option value="no_highway">🛣️ 不走高速</option>
                            <option value="avoid_jam">🚦 躲避拥堵</option>
                            <option value="fastest_free">⚡💰 时间最短且免费</option>
                            <option value="shortest_free">📏💰 距离最短且免费</option>
                        </select>
                    </div>
                    <div style="display: flex; align-items: end; gap: 10px; flex-wrap: wrap;">
                        <button onclick="calculateRoute()" class="btn btn-primary" id="calcBtn">
                            <span>🧮</span>计算路程
                        </button>
                        <button onclick="showAddLocation()" class="btn btn-success">
                            <span>➕</span>添加地点
                        </button>
                    </div>
                </div>

                <div id="routeResult" style="display: none;"></div>

                <div class="form-group">
                    <label>📝 工作内容</label>
                    <textarea id="workNotes" class="form-control" rows="3" placeholder="请描述本次工作的具体内容..."></textarea>
                </div>

                <div style="margin-top: 30px;">
                    <button onclick="submitTimesheet()" class="btn btn-primary" id="submitBtn">
                        <span>✅</span>提交工时记录
                    </button>
                    <button onclick="loadHistory()" class="btn btn-info">
                        <span>📊</span>查看历史
                    </button>
                    <button onclick="resetToUserPanel()" class="btn btn-secondary">
                        <span>🔄</span>切换用户
                    </button>
                </div>
            </div>

            <!-- 历史记录面板 -->
            <div id="historyPanel" style="display: none;">
                <h3>📊 工时历史记录</h3>
                <div class="table-container">
                    <table class="table" id="historyTable">
                        <thead>
                            <tr>
                                <th>日期</th>
                                <th>路线</th>
                                <th>交通方式</th>
                                <th>距离</th>
                                <th>时间</th>
                                <th>工作时长</th>
                                <th>数据源</th>
                                <th>备注</th>
                            </tr>
                        </thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- 门店管理面板 -->
        <div id="stores-panel" class="card">
            <h3>🏪 门店信息管理</h3>
            
            <!-- 导入门店信息 -->
            <div style="margin-bottom: 30px;">
                <h4>📥 导入门店信息</h4>
                <div class="file-upload" onclick="document.getElementById('storeFile').click();" 
                     ondrop="handleDrop(event)" ondragover="handleDragOver(event)" ondragleave="handleDragLeave(event)">
                    <input type="file" id="storeFile" accept=".csv,.xlsx,.json" style="display: none;" onchange="handleFileSelect(event)">
                    <div style="font-size: 48px; margin-bottom: 15px;">📄</div>
                    <div style="font-size: 18px; font-weight: 600; margin-bottom: 10px;">点击选择或拖拽文件</div>
                    <div style="color: #6c757d;">支持 CSV、Excel、JSON 格式</div>
                    <div style="color: #6c757d; font-size: 14px; margin-top: 10px;">
                        文件格式：门店编码,门店名称,门店城市,经度,纬度,地址
                    </div>
                </div>
                <button onclick="downloadTemplate()" class="btn btn-info">
                    <span>📥</span>下载模板文件
                </button>
            </div>

            <!-- 门店列表 -->
            <div>
                <h4>🏪 门店列表</h4>
                
                <!-- 搜索和统计 -->
                <div style="margin-bottom: 20px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 15px;">
                    <div style="display: flex; align-items: center; gap: 10px;">
                        <input type="text" id="storeSearch" class="form-control" placeholder="搜索门店编码、名称或城市..." style="width: 300px;">
                        <button onclick="searchStores()" class="btn btn-primary">🔍 搜索</button>
                        <button onclick="resetSearch()" class="btn btn-secondary">🔄 重置</button>
                    </div>
                    <div id="storeStats" style="color: #666; font-weight: 500;">
                        加载中...
                    </div>
                </div>
                
                <div class="table-container">
                    <table class="table" id="storesTable">
                        <thead>
                            <tr>
                                <th>门店编码</th>
                                <th>门店名称</th>
                                <th>所在城市</th>
                                <th>经度</th>
                                <th>纬度</th>
                                <th>地址</th>
                                <th>创建时间</th>
                                <th>操作</th>
                            </tr>
                        </thead>
                        <tbody id="storesTableBody">
                            <tr>
                                <td colspan="8" style="text-align: center; padding: 40px; color: #666;">
                                    <div class="loading"></div>
                                    <div style="margin-top: 10px;">加载门店数据中...</div>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <!-- 分页控件 -->
                <div id="storePagination" style="margin-top: 20px; text-align: center; display: none;">
                    <div style="display: inline-flex; align-items: center; gap: 10px; flex-wrap: wrap;">
                        <button onclick="loadStores(currentStoresPage - 1)" class="btn btn-secondary" id="prevPageBtn" disabled>
                            ← 上一页
                        </button>
                        <span id="pageInfo" style="margin: 0 15px; font-weight: 500;">
                            第 1 页，共 1 页
                        </span>
                        <button onclick="loadStores(currentStoresPage + 1)" class="btn btn-secondary" id="nextPageBtn" disabled>
                            下一页 →
                        </button>
                        <select id="pageSizeSelect" class="form-control" onchange="changePageSize()" style="width: 120px; margin-left: 20px;">
                            <option value="20">20/页</option>
                            <option value="50" selected>50/页</option>
                            <option value="100">100/页</option>
                        </select>
                    </div>
                </div>
            </div>
        </div>

        <!-- 数据导出面板 -->
        <div id="export-panel" class="card">
            <h3>📊 数据导出</h3>
            
            <div class="form-grid">
                <div class="form-group">
                    <label>选择用户</label>
                    <select id="exportUserSelect" class="form-control">
                        <option value="">全部用户</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>开始日期</label>
                    <input type="date" id="exportStartDate" class="form-control">
                </div>
                <div class="form-group">
                    <label>结束日期</label>
                    <input type="date" id="exportEndDate" class="form-control">
                </div>
            </div>

            <div style="margin-top: 30px;">
                <button onclick="exportExcel()" class="btn btn-success" id="exportBtn">
                    <span>📊</span>导出Excel工时表
                </button>
                <button onclick="exportStoresExcel()" class="btn btn-info">
                    <span>🏪</span>导出门店信息
                </button>
                <button onclick="exportJSON()" class="btn btn-warning">
                    <span>📄</span>导出JSON数据
                </button>
            </div>
        </div>
    </div>

    <script>
        let currentUser = null;
        let users = [];
        let locations = [];
        let stores = [];
        let currentTab = 'timesheet';
        let currentStoresPage = 1;
        let currentStoresSearch = '';
        let storesPerPage = 50;

        // 加载用户信息
        async function loadUserProfile() {
            try {
                const response = await fetch('/api/profile');
                if (response.ok) {
                    const user = await response.json();
                    document.getElementById('currentUser').textContent = `${user.name} (${user.department})`;
                } else if (response.status === 401) {
                    // 未登录，跳转到登录页面
                    window.location.href = '/login';
                } else {
                    document.getElementById('currentUser').textContent = '未知用户';
                }
            } catch (error) {
                console.error('加载用户信息失败:', error);
                document.getElementById('currentUser').textContent = '加载失败';
            }
        }

        // 退出登录
        async function logout() {
            try {
                const response = await fetch('/api/logout', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    }
                });
                
                if (response.ok) {
                    alert('已退出登录');
                    window.location.href = '/login';
                } else {
                    alert('退出失败，请重试');
                }
            } catch (error) {
                console.error('退出登录失败:', error);
                alert('网络错误，请重试');
            }
        }

        // 页面初始化
        document.addEventListener('DOMContentLoaded', function() {
            console.log('🚀 系统初始化中...');
            loadUserProfile();
            loadUsers();
            loadLocations();
            
            // 初始化时不加载门店，在需要时按需加载
            console.log('📋 初始化完成，门店数据将按需加载');
            
            // 初始化路线策略显示状态
            onTransportModeChange();
            
            document.getElementById('workDate').value = new Date().toISOString().split('T')[0];
            
            // 设置导出日期默认值为本月
            const now = new Date();
            const firstDay = new Date(now.getFullYear(), now.getMonth(), 1);
            const lastDay = new Date(now.getFullYear(), now.getMonth() + 1, 0);
            document.getElementById('exportStartDate').value = firstDay.toISOString().split('T')[0];
            document.getElementById('exportEndDate').value = lastDay.toISOString().split('T')[0];
        });

        function switchTab(tabName) {
            // 更新导航状态
            document.querySelectorAll('.nav-tab').forEach(tab => tab.classList.remove('active'));
            event.target.classList.add('active');
            
            // 显示对应面板
            document.querySelectorAll('.card').forEach(panel => panel.classList.remove('active'));
            document.getElementById(tabName + '-panel').classList.add('active');
            
            currentTab = tabName;
            
            // 根据标签页加载对应数据
            if (tabName === 'stores') {
                console.log('🏪 切换到门店管理标签');
                loadStores(1, ''); // 加载门店列表（分页）
            } else if (tabName === 'export') {
                console.log('📊 切换到数据导出标签');
                loadUsersForExport();
            } else if (tabName === 'timesheet') {
                console.log('⏰ 切换到工时管理标签');
                // 延迟加载门店选择器，避免阻塞UI
                setTimeout(() => {
                    if (currentTab === 'timesheet') {
                        updateStoreSelectors();
                    }
                }, 100);
            }
        }

        function showMessage(text, type = 'success') {
            const messageDiv = document.getElementById('message');
            let icon = '✅';
            if (type === 'error') icon = '❌';
            else if (type === 'info') icon = 'ℹ️';
            
            messageDiv.innerHTML = `<div class="message ${type}"><span>${icon}</span><span>${text}</span></div>`;
            setTimeout(() => { messageDiv.innerHTML = ''; }, type === 'info' ? 6000 : 4000);
        }

        async function loadUsers() {
            try {
                const response = await fetch('/api/users');
                users = await response.json();
                const select = document.getElementById('userSelect');
                select.innerHTML = '<option value="">请选择用户...</option>';
                users.forEach(user => {
                    select.innerHTML += `<option value="${user.id}">${user.name} (${user.department})</option>`;
                });
                console.log(`✅ 加载了 ${users.length} 个用户`);
            } catch (error) {
                showMessage('加载用户列表失败', 'error');
                console.error('❌ 加载用户失败:', error);
            }
        }

        async function loadUsersForExport() {
            try {
                const response = await fetch('/api/users');
                const users = await response.json();
                const select = document.getElementById('exportUserSelect');
                select.innerHTML = '<option value="">全部用户</option>';
                users.forEach(user => {
                    select.innerHTML += `<option value="${user.id}">${user.name} (${user.department})</option>`;
                });
            } catch (error) {
                console.error('加载导出用户列表失败:', error);
            }
        }

        async function loadLocations() {
            try {
                const response = await fetch('/api/locations');
                locations = await response.json();
                
                const selects = ['fromLocation', 'toLocation'];
                selects.forEach(selectId => {
                    const select = document.getElementById(selectId);
                    const placeholder = selectId === 'fromLocation' ? '选择出发地点...' : '选择目的地点...';
                    select.innerHTML = `<option value="">${placeholder}</option>`;
                    locations.forEach(location => {
                        select.innerHTML += `<option value="${location.id}">${location.name}</option>`;
                    });
                });
                console.log(`✅ 加载了 ${locations.length} 个地点`);
            } catch (error) {
                showMessage('加载地点列表失败', 'error');
                console.error('❌ 加载地点失败:', error);
            }
        }

        async function loadStores(page = 1, search = '') {
            try {
                const params = new URLSearchParams({
                    page: page,
                    per_page: storesPerPage
                });
                
                if (search) {
                    params.append('search', search);
                }
                
                const response = await fetch(`/api/stores?${params.toString()}`);
                const data = await response.json();
                
                stores = data.stores;
                const pagination = data.pagination;
                
                // 更新页面状态
                currentStoresPage = page;
                currentStoresSearch = search;
                
                // 更新门店选择器（只在第一次加载或搜索时更新）
                if (page === 1) {
                    await updateStoreSelectors();
                }
                
                // 更新门店表格
                updateStoresTable(stores);
                
                // 更新分页控件
                updateStorePagination(pagination);
                
                // 更新统计信息
                updateStoreStats(pagination);
                
                console.log(`✅ 加载了第 ${page} 页门店数据，共 ${stores.length} 个`);
            } catch (error) {
                showMessage('加载门店列表失败', 'error');
                console.error('❌ 加载门店失败:', error);
            }
        }

        async function updateStoreSelectors() {
            try {
                // 只在工时管理标签需要时才加载门店选择器
                if (currentTab !== 'timesheet') {
                    return;
                }
                
                // 分批加载门店数据，避免一次性加载过多
                console.log('🔄 正在更新门店选择器...');
                const response = await fetch('/api/stores?per_page=1000'); // 限制数量
                const data = await response.json();
                const allStores = data.stores;
                
                const selects = ['fromStore', 'toStore'];
                selects.forEach(selectId => {
                    const select = document.getElementById(selectId);
                    if (!select) return;
                    
                    const placeholder = selectId === 'fromStore' ? '选择出发门店...' : '选择目的门店...';
                    select.innerHTML = `<option value="">${placeholder}</option>`;
                    
                    // 分批添加选项，避免阻塞UI
                    const batchSize = 100;
                    let index = 0;
                    
                    function addBatch() {
                        const endIndex = Math.min(index + batchSize, allStores.length);
                        for (let i = index; i < endIndex; i++) {
                            const store = allStores[i];
                            const option = document.createElement('option');
                            option.value = store.id;
                            option.textContent = `${store.store_code} - ${store.store_name} (${store.store_city})`;
                            select.appendChild(option);
                        }
                        
                        index = endIndex;
                        if (index < allStores.length) {
                            // 使用 setTimeout 避免阻塞UI
                            setTimeout(addBatch, 10);
                        } else {
                            console.log(`✅ 门店选择器更新完成，共 ${allStores.length} 个门店`);
                        }
                    }
                    
                    addBatch();
                });
                
                // 如果门店数量太多，显示提示
                if (data.pagination && data.pagination.total > 1000) {
                    showMessage(`门店数据较多(${data.pagination.total}个)，选择器只显示前1000个门店`, 'info');
                }
                
            } catch (error) {
                console.error('更新门店选择器失败:', error);
                showMessage('门店选择器加载失败，请刷新页面重试', 'error');
            }
        }

        function updateStoresTable(stores) {
            const tbody = document.getElementById('storesTableBody');
            if (stores.length === 0) {
                tbody.innerHTML = `
                    <tr>
                        <td colspan="8" style="text-align: center; padding: 40px; color: #666;">
                            ${currentStoresSearch ? '🔍 没有找到匹配的门店' : '📭 暂无门店数据'}
                        </td>
                    </tr>
                `;
                return;
            }
            
            tbody.innerHTML = '';
            stores.forEach(store => {
                tbody.innerHTML += `
                    <tr>
                        <td>${store.store_code}</td>
                        <td>${store.store_name}</td>
                        <td>${store.store_city}</td>
                        <td>${store.longitude}</td>
                        <td>${store.latitude}</td>
                        <td>${store.address || '-'}</td>
                        <td>${store.created_at ? store.created_at.split('T')[0] : '-'}</td>
                        <td>
                            <button onclick="deleteStore(${store.id})" class="btn btn-secondary" style="padding: 6px 12px; font-size: 14px;">
                                🗑️ 删除
                            </button>
                        </td>
                    </tr>
                `;
            });
        }

        function updateStorePagination(pagination) {
            const paginationDiv = document.getElementById('storePagination');
            const prevBtn = document.getElementById('prevPageBtn');
            const nextBtn = document.getElementById('nextPageBtn');
            const pageInfo = document.getElementById('pageInfo');
            
            if (pagination.pages <= 1) {
                paginationDiv.style.display = 'none';
                return;
            }
            
            paginationDiv.style.display = 'block';
            
            // 更新按钮状态
            prevBtn.disabled = pagination.page <= 1;
            nextBtn.disabled = pagination.page >= pagination.pages;
            
            // 更新页面信息
            pageInfo.textContent = `第 ${pagination.page} 页，共 ${pagination.pages} 页`;
        }

        function updateStoreStats(pagination) {
            const statsDiv = document.getElementById('storeStats');
            const start = (pagination.page - 1) * pagination.per_page + 1;
            const end = Math.min(pagination.page * pagination.per_page, pagination.total);
            
            let statsText = `显示第 ${start}-${end} 条，共 ${pagination.total} 个门店`;
            if (currentStoresSearch) {
                statsText += ` (搜索: "${currentStoresSearch}")`;
            }
            
            statsDiv.textContent = statsText;
        }

        function searchStores() {
            const searchInput = document.getElementById('storeSearch');
            const searchTerm = searchInput.value.trim();
            loadStores(1, searchTerm);
        }

        function resetSearch() {
            document.getElementById('storeSearch').value = '';
            loadStores(1, '');
        }

        function changePageSize() {
            const select = document.getElementById('pageSizeSelect');
            storesPerPage = parseInt(select.value);
            loadStores(1, currentStoresSearch);
        }

        // 回车搜索
        document.addEventListener('DOMContentLoaded', function() {
            const searchInput = document.getElementById('storeSearch');
            if (searchInput) {
                searchInput.addEventListener('keypress', function(e) {
                    if (e.key === 'Enter') {
                        searchStores();
                    }
                });
            }
        });

        function toggleLocationOptions(type) {
            const typeSelect = document.getElementById(type + 'Type');
            const locationSelect = document.getElementById(type + 'Location');
            const storeContainer = document.getElementById(type + 'StoreContainer');
            const storeInput = document.getElementById(type + 'StoreCode');
            const storeName = document.getElementById(type + 'StoreName');
            const label = document.getElementById(type + 'Label');
            
            if (typeSelect.value === 'store') {
                locationSelect.style.display = 'none';
                storeContainer.style.display = 'block';
                label.textContent = type === 'from' ? '🏪 出发门店编码' : '🎯 目的门店编码';
            } else {
                locationSelect.style.display = 'block';
                storeContainer.style.display = 'none';
                label.textContent = type === 'from' ? '📍 出发地点' : '🎯 目的地点';
            }
            
            // 重置选择
            locationSelect.value = '';
            storeInput.value = '';
            storeName.textContent = '';
            storeName.className = 'store-name-display empty';
            
            autoCalculate();
        }

        // 门店编码输入处理函数
        async function onStoreCodeInput(type) {
            const storeInput = document.getElementById(type + 'StoreCode');
            const storeName = document.getElementById(type + 'StoreName');
            const storeCode = storeInput.value.trim();
            
            if (!storeCode) {
                storeName.textContent = '';
                storeName.className = 'store-name-display empty';
                autoCalculate();
                return;
            }
            
            // 显示加载状态
            storeName.textContent = '🔍 查找门店中...';
            storeName.className = 'store-name-display';
            
            try {
                // 调用新的API端点根据门店编码获取门店信息
                const response = await fetch(`/api/stores/by-code/${storeCode}`);
                const data = await response.json();
                
                if (data.success) {
                    const store = data.store;
                    storeName.innerHTML = `
                        <div style="color: #28a745; display: flex; align-items: center; gap: 8px; flex-wrap: wrap;">
                            <span>✅</span>
                            <div>
                                <div style="font-weight: 600; font-size: 14px;">${store.store_name}</div>
                                <div style="font-size: 12px; color: #6c757d;">
                                    ${store.store_city} • ${store.address}
                                </div>
                            </div>
                        </div>
                    `;
                    storeName.className = 'store-name-display found';
                } else {
                    storeName.innerHTML = `
                        <div style="color: #dc3545; display: flex; align-items: center; gap: 8px;">
                            <span>❌</span>
                            <span>门店编码不存在: ${storeCode}</span>
                        </div>
                    `;
                    storeName.className = 'store-name-display not-found';
                }
            } catch (error) {
                console.error('查找门店失败:', error);
                storeName.innerHTML = `
                    <div style="color: #ffc107; display: flex; align-items: center; gap: 8px;">
                        <span>⚠️</span>
                        <span>查找失败，请稍后重试</span>
                    </div>
                `;
                storeName.className = 'store-name-display not-found';
            }
            
            autoCalculate();
        }

        function selectUser() {
            const userId = document.getElementById('userSelect').value;
            if (!userId) {
                showMessage('请先选择一个用户', 'error');
                return;
            }
            
            currentUser = users.find(u => u.id == userId);
            if (currentUser) {
                document.getElementById('workTitle').textContent = `📝 ${currentUser.name} 的工时录入`;
                document.getElementById('userPanel').style.display = 'none';
                document.getElementById('workPanel').style.display = 'block';
                document.getElementById('statsPanel').style.display = 'none';
                showMessage(`欢迎，${currentUser.name}！开始记录工时吧`);
            }
        }

        function resetToUserPanel() {
            currentUser = null;
            document.getElementById('userPanel').style.display = 'block';
            document.getElementById('workPanel').style.display = 'none';
            document.getElementById('historyPanel').style.display = 'none';
            document.getElementById('statsPanel').style.display = 'none';
        }

        function autoCalculate() {
            const fromType = document.getElementById('fromType').value;
            const toType = document.getElementById('toType').value;
            
            let fromId, toId;
            
            if (fromType === 'store') {
                fromId = document.getElementById('fromStore').value;
            } else {
                fromId = document.getElementById('fromLocation').value;
            }
            
            if (toType === 'store') {
                toId = document.getElementById('toStore').value;
            } else {
                toId = document.getElementById('toLocation').value;
            }
            
            if (fromId && toId && fromId !== toId) {
                setTimeout(calculateRoute, 300);
            }
        }
        
        function onTransportModeChange() {
            const transportMode = document.getElementById('transportMode').value;
            const routeStrategyGroup = document.getElementById('routeStrategyGroup');
            
            // 只有自驾和打车可以选择路线策略
            if (transportMode === '自驾' || transportMode === '打车') {
                routeStrategyGroup.style.display = 'block';
            } else {
                routeStrategyGroup.style.display = 'none';
            }
            
            autoCalculate();
        }

        async function calculateRoute() {
            const fromType = document.getElementById('fromType').value;
            const toType = document.getElementById('toType').value;
            const transportMode = document.getElementById('transportMode').value;
            const routeStrategy = document.getElementById('routeStrategy').value;
            
            let fromId, toId;
            
            if (fromType === 'store') {
                fromId = document.getElementById('fromStoreCode').value.trim();
            } else {
                fromId = document.getElementById('fromLocation').value;
            }
            
            if (toType === 'store') {
                toId = document.getElementById('toStoreCode').value.trim();
            } else {
                toId = document.getElementById('toLocation').value;
            }
            
            if (!fromId || !toId) {
                showMessage('请输入出发地和目的地', 'error');
                return;
            }
            
            if (fromId === toId && fromType === toType) {
                showMessage('出发地和目的地不能相同', 'error');
                return;
            }
            
            const calcBtn = document.getElementById('calcBtn');
            calcBtn.innerHTML = '<span class="loading"></span><span>计算中...</span>';
            calcBtn.disabled = true;
            
            try {
                const requestData = {
                    transport_mode: transportMode,
                    route_strategy: routeStrategy
                };
                
                if (fromType === 'store') {
                    requestData.store_from_code = fromId;
                } else {
                    requestData.location_from_id = parseInt(fromId);
                }
                
                if (toType === 'store') {
                    requestData.store_to_code = toId;
                } else {
                    requestData.location_to_id = parseInt(toId);
                }
                
                const response = await fetch('/api/calculate-route', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(requestData)
                });
                
                const result = await response.json();
                if (response.ok) {
                    displayRouteResult(result);
                    showMessage('路程计算完成');
                } else {
                    showMessage(`计算失败：${result.error}`, 'error');
                }
            } catch (error) {
                showMessage('网络错误，请重试', 'error');
                console.error('❌ 路程计算失败:', error);
            } finally {
                calcBtn.innerHTML = '<span>🧮</span><span>计算路程</span>';
                calcBtn.disabled = false;
            }
        }

        function displayRouteResult(result) {
            const resultDiv = document.getElementById('routeResult');
            const badgeClass = result.api_used ? 'badge-api' : 'badge-fallback';
            const badgeText = result.api_used ? '🗺️ 高德API' : '📐 智能估算';
            
            resultDiv.innerHTML = `
                <div class="route-result">
                    <div class="route-header">
                        📍 ${result.from_name} → ${result.to_name}
                        <span class="badge ${badgeClass}" style="float: right;">${badgeText}</span>
                    </div>
                    <div class="route-details">
                        <div class="route-item">
                            <div class="route-value">${result.distance}</div>
                            <div class="route-label">距离 (km)</div>
                        </div>
                        <div class="route-item">
                            <div class="route-value">${result.travel_time}</div>
                            <div class="route-label">时间 (h)</div>
                        </div>
                        <div class="route-item">
                            <div class="route-value">${result.transport_mode}</div>
                            <div class="route-label">交通方式</div>
                        </div>
                        <div class="route-item">
                            <div class="route-value">${(result.distance / result.travel_time).toFixed(1)}</div>
                            <div class="route-label">平均速度 (km/h)</div>
                        </div>
                    </div>
                </div>
            `;
            resultDiv.style.display = 'block';
        }

        async function submitTimesheet() {
            if (!currentUser) {
                showMessage('请先选择用户', 'error');
                return;
            }
            
            const fromType = document.getElementById('fromType').value;
            const toType = document.getElementById('toType').value;
            
            const formData = {
                user_id: currentUser.id,
                date: document.getElementById('workDate').value,
                transport_mode: document.getElementById('transportMode').value,
                work_hours: parseFloat(document.getElementById('workHours').value),
                notes: document.getElementById('workNotes').value.trim()
            };
            
            // 根据类型设置起止点
            if (fromType === 'store') {
                formData.store_from_code = document.getElementById('fromStoreCode').value.trim() || null;
            } else {
                formData.location_from_id = parseInt(document.getElementById('fromLocation').value) || null;
            }
            
            if (toType === 'store') {
                formData.store_to_code = document.getElementById('toStoreCode').value.trim() || null;
            } else {
                formData.location_to_id = parseInt(document.getElementById('toLocation').value) || null;
            }
            
            if (!formData.date || !formData.work_hours) {
                showMessage('请填写工作日期和工作时长', 'error');
                return;
            }
            
            const submitBtn = document.getElementById('submitBtn');
            submitBtn.innerHTML = '<span class="loading"></span><span>提交中...</span>';
            submitBtn.disabled = true;
            
            try {
                const response = await fetch('/api/timesheet', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(formData)
                });
                
                if (response.ok) {
                    showMessage('工时记录提交成功！');
                    resetForm();
                } else {
                    const error = await response.json();
                    showMessage(`提交失败：${error.error}`, 'error');
                }
            } catch (error) {
                showMessage('网络错误，请重试', 'error');
                console.error('❌ 提交失败:', error);
            } finally {
                submitBtn.innerHTML = '<span>✅</span><span>提交工时记录</span>';
                submitBtn.disabled = false;
            }
        }

        function resetForm() {
            document.getElementById('workDate').value = new Date().toISOString().split('T')[0];
            document.getElementById('workHours').value = '8';
            document.getElementById('fromLocation').value = '';
            document.getElementById('toLocation').value = '';
            document.getElementById('fromStore').value = '';
            document.getElementById('toStore').value = '';
            document.getElementById('transportMode').value = '自驾';
            document.getElementById('workNotes').value = '';
            document.getElementById('routeResult').style.display = 'none';
        }

        async function loadHistory() {
            if (!currentUser) return;
            
            try {
                const response = await fetch(`/api/timesheet?user_id=${currentUser.id}`);
                const records = await response.json();
                
                const tbody = document.querySelector('#historyTable tbody');
                tbody.innerHTML = '';
                
                records.forEach(record => {
                    let route = '本地工作';
                    if (record.from_name && record.to_name) {
                        route = `${record.from_name} → ${record.to_name}`;
                    }
                    
                    const distance = record.distance ? `${record.distance} km` : '-';
                    const time = record.travel_time ? `${record.travel_time} h` : '-';
                    const apiSource = record.api_used ? 
                        '<span class="badge badge-api">🗺️ API</span>' : 
                        '<span class="badge badge-fallback">📐 估算</span>';
                    
                    tbody.innerHTML += `
                        <tr>
                            <td>${record.date}</td>
                            <td>${route}</td>
                            <td>${record.transport_mode}</td>
                            <td>${distance}</td>
                            <td>${time}</td>
                            <td>${record.work_hours} h</td>
                            <td>${apiSource}</td>
                            <td>${record.notes || '-'}</td>
                        </tr>
                    `;
                });
                
                document.getElementById('historyPanel').style.display = 'block';
                showMessage(`加载了 ${records.length} 条历史记录`);
            } catch (error) {
                showMessage('加载历史记录失败', 'error');
                console.error('❌ 加载历史失败:', error);
            }
        }

        async function showUserStats() {
            const userId = document.getElementById('userSelect').value;
            if (!userId) {
                showMessage('请先选择用户', 'error');
                return;
            }
            
            try {
                const response = await fetch(`/api/statistics/${userId}`);
                const stats = await response.json();
                
                const statsDiv = document.getElementById('statsPanel');
                statsDiv.innerHTML = `
                    <div class="stat-card">
                        <div class="stat-number">${stats.total_work_hours}</div>
                        <div class="stat-label">总工时 (h)</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">${stats.work_days}</div>
                        <div class="stat-label">工作天数</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">${stats.total_distance}</div>
                        <div class="stat-label">总距离 (km)</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">${stats.total_travel_time}</div>
                        <div class="stat-label">行程时间 (h)</div>
                    </div>
                `;
                statsDiv.style.display = 'grid';
                
                const user = users.find(u => u.id == userId);
                showMessage(`${user.name} 的统计数据已加载`);
            } catch (error) {
                showMessage('加载统计数据失败', 'error');
                console.error('❌ 加载统计失败:', error);
            }
        }

        function showAddLocation() {
            showMessage('添加地点功能，请切换到门店管理面板', 'error');
        }

        // 门店管理相关函数
        function handleDragOver(e) {
            e.preventDefault();
            e.currentTarget.classList.add('dragover');
        }

        function handleDragLeave(e) {
            e.currentTarget.classList.remove('dragover');
        }

        function handleDrop(e) {
            e.preventDefault();
            e.currentTarget.classList.remove('dragover');
            const files = e.dataTransfer.files;
            if (files.length > 0) {
                processStoreFile(files[0]);
            }
        }

        function handleFileSelect(e) {
            const file = e.target.files[0];
            if (file) {
                processStoreFile(file);
            }
        }

        async function processStoreFile(file) {
            const formData = new FormData();
            formData.append('file', file);
            
            try {
                showMessage('正在处理文件...', 'success');
                const response = await fetch('/api/stores/import', {
                    method: 'POST',
                    body: formData
                });
                
                const result = await response.json();
                if (response.ok) {
                    showMessage(`导入成功！新增 ${result.imported} 个门店`);
                    loadStores();
                } else {
                    showMessage(`导入失败：${result.error}`, 'error');
                }
            } catch (error) {
                showMessage('文件处理失败', 'error');
                console.error('导入文件失败:', error);
            }
        }

        async function deleteStore(storeId) {
            if (confirm('确认删除这个门店吗？')) {
                try {
                    const response = await fetch(`/api/stores/${storeId}`, {
                        method: 'DELETE'
                    });
                    
                    if (response.ok) {
                        showMessage('门店删除成功');
                        loadStores();
                    } else {
                        showMessage('删除失败', 'error');
                    }
                } catch (error) {
                    showMessage('网络错误', 'error');
                }
            }
        }

        function downloadTemplate() {
            const template = "门店编码,门店名称,门店城市,经度,纬度,地址\\n" +
                           "HZ001,杭州西湖店,杭州市,120.1552,30.2741,杭州市西湖区文三路100号\\n" +
                           "SH001,上海徐汇店,上海市,121.4737,31.2304,上海市徐汇区漕溪北路88号";
            
            const blob = new Blob([template], { type: 'text/csv;charset=utf-8;' });
            const link = document.createElement('a');
            link.href = URL.createObjectURL(blob);
            link.download = '门店信息导入模板.csv';
            link.click();
        }

        // 数据导出相关函数
        async function exportExcel() {
            const userId = document.getElementById('exportUserSelect').value;
            const startDate = document.getElementById('exportStartDate').value;
            const endDate = document.getElementById('exportEndDate').value;
            
            const exportBtn = document.getElementById('exportBtn');
            exportBtn.innerHTML = '<span class="loading"></span><span>导出中...</span>';
            exportBtn.disabled = true;
            
            try {
                const params = new URLSearchParams();
                if (userId) params.append('user_id', userId);
                if (startDate) params.append('start_date', startDate);
                if (endDate) params.append('end_date', endDate);
                
                const response = await fetch(`/api/export/timesheet?${params.toString()}`);
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `工时表_${startDate}_${endDate}.xlsx`;
                    a.click();
                    window.URL.revokeObjectURL(url);
                    showMessage('Excel文件导出成功');
                } else {
                    showMessage('导出失败', 'error');
                }
            } catch (error) {
                showMessage('网络错误，请重试', 'error');
                console.error('导出失败:', error);
            } finally {
                exportBtn.innerHTML = '<span>📊</span><span>导出Excel工时表</span>';
                exportBtn.disabled = false;
            }
        }

        async function exportStoresExcel() {
            try {
                const response = await fetch('/api/export/stores');
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = '门店信息.xlsx';
                    a.click();
                    window.URL.revokeObjectURL(url);
                    showMessage('门店Excel文件导出成功');
                } else {
                    showMessage('导出失败', 'error');
                }
            } catch (error) {
                showMessage('网络错误，请重试', 'error');
            }
        }

        async function exportJSON() {
            const userId = document.getElementById('exportUserSelect').value;
            const startDate = document.getElementById('exportStartDate').value;
            const endDate = document.getElementById('exportEndDate').value;
            
            try {
                const params = new URLSearchParams();
                if (userId) params.append('user_id', userId);
                if (startDate) params.append('start_date', startDate);
                if (endDate) params.append('end_date', endDate);
                
                const response = await fetch(`/api/export/json?${params.toString()}`);
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `工时数据_${startDate}_${endDate}.json`;
                    a.click();
                    window.URL.revokeObjectURL(url);
                    showMessage('JSON文件导出成功');
                } else {
                    showMessage('导出失败', 'error');
                }
            } catch (error) {
                showMessage('网络错误，请重试', 'error');
            }
        }
    </script>
</body>
</html>
    '''

# 健康检查API
@app.route('/api/health')
def health_check():
    """健康检查端点"""
    try:
        # 检查数据库连接
        conn = sqlite3.connect('enhanced_timesheet.db')
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        conn.close()
        
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'version': '2.0.0',
            'services': {
                'database': 'ok',
                'auth': 'ok'
            }
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'timestamp': datetime.now().isoformat(),
            'error': str(e)
        }), 500

# 用户认证API
@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    
    required_fields = ['name', 'group', 'password', 'role']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'缺少必要字段: {field}'}), 400
    
    # 验证密码长度
    if len(data['password']) < 6:
        return jsonify({'error': '密码长度至少6位'}), 400
    
    # 验证组别是否有效
    valid_groups = ['稽核一组', '稽核二组', '稽核三组', '稽核四组']
    if data['group'] not in valid_groups:
        return jsonify({'error': '无效的组别'}), 400
    
    # 验证角色是否有效
    valid_roles = ['普通用户', '管理员']
    if data['role'] not in valid_roles:
        return jsonify({'error': '无效的角色'}), 400
    
    # 密码加密
    password_hash = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt())
    
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    try:
        # 检查是否已存在相同姓名和组别的用户
        cursor.execute('''
            SELECT id FROM users WHERE name = ? AND group_name = ?
        ''', (data['name'], data['group']))
        
        existing_user = cursor.fetchone()
        if existing_user:
            conn.close()
            return jsonify({'error': '该姓名在此组别中已存在'}), 400
        
        # 生成唯一的用户名（姓名+组别+时间戳）
        import time
        username = f"{data['name']}_{data['group']}_{int(time.time())}"
        
        # 插入新用户
        cursor.execute('''
            INSERT INTO users (username, name, group_name, department, position, email, password_hash)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            username,
            data['name'],
            data['group'],
            data['group'],  # 部门设为组别
            data['role'],   # 职位设为用户选择的角色
            f"{username}@company.com",  # 生成默认邮箱
            password_hash.decode('utf-8')
        ))
        
        user_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '注册成功',
            'user_id': user_id,
            'username': username
        }), 201
        
    except sqlite3.Error as e:
        conn.close()
        return jsonify({'error': f'数据库错误: {str(e)}'}), 500

# 获取用户列表API
@app.route('/api/users', methods=['GET'])
def get_users():
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT id, name, group_name FROM users 
            WHERE is_active = 1 
            ORDER BY group_name, name
        ''')
        
        users = []
        for row in cursor.fetchall():
            users.append({
                'id': row[0],
                'name': row[1],
                'group_name': row[2] or '未分组'
            })
        
        conn.close()
        return jsonify(users)
        
    except sqlite3.Error as e:
        conn.close()
        return jsonify({'error': f'数据库错误: {str(e)}'}), 500

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    
    required_fields = ['name', 'group', 'password']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'缺少必要字段: {field}'}), 400
    
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT id, name, department, position, group_name, is_active, password_hash
            FROM users WHERE name = ? AND group_name = ?
        ''', (data['name'], data['group']))
        
        user = cursor.fetchone()
        conn.close()
        
        if not user:
            return jsonify({'error': '用户不存在或组别不匹配'}), 401
        
        if not user[5]:  # is_active
            return jsonify({'error': '账户已被禁用'}), 401
        
        # 验证密码
        if not user[6] or not bcrypt.checkpw(data['password'].encode('utf-8'), user[6].encode('utf-8')):
            return jsonify({'error': '密码错误'}), 401
        
        # 设置session
        session['user_id'] = user[0]
        session['name'] = user[1]
        session['department'] = user[2]
        session['position'] = user[3]
        session['group_name'] = user[4]
        
        return jsonify({
            'success': True,
            'message': '登录成功',
            'user': {
                'id': user[0],
                'name': user[1],
                'department': user[2],
                'position': user[3],
                'group_name': user[4]
            }
        })
        
    except sqlite3.Error as e:
        conn.close()
        return jsonify({'error': f'数据库错误: {str(e)}'}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True, 'message': '已退出登录'})

@app.route('/api/user/info', methods=['GET'])
@login_required
def get_user_info():
    """获取当前登录用户信息"""
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT id, name, department, position, group_name, email, is_active
            FROM users WHERE id = ?
        ''', (session['user_id'],))
        
        user = cursor.fetchone()
        conn.close()
        
        if not user:
            return jsonify({'error': '用户不存在'}), 404
        
        return jsonify({
            'id': user[0],
            'name': user[1],
            'department': user[2],
            'position': user[3],
            'group_name': user[4] or '未分组',
            'email': user[5],
            'is_active': user[6]
        })
        
    except sqlite3.Error as e:
        conn.close()
        return jsonify({'error': f'数据库错误: {str(e)}'}), 500

@app.route('/api/stores/import', methods=['POST'])
@login_required
def import_stores():
    """导入门店信息 - 仅管理员可用"""
    # 检查用户权限
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    cursor.execute("SELECT position FROM users WHERE id = ?", (session['user_id'],))
    user = cursor.fetchone()
    
    if not user or user[0] != '管理员':
        conn.close()
        return jsonify({'error': '权限不足，仅管理员可导入门店信息'}), 403
    
    if 'file' not in request.files:
        conn.close()
        return jsonify({'error': '未选择文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        conn.close()
        return jsonify({'error': '未选择文件'}), 400
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        conn.close()
        return jsonify({'error': '文件格式不支持，请上传Excel文件'}), 400
    
    try:
        import pandas as pd
        
        # 读取Excel文件
        df = pd.read_excel(file)
        
        # 检查必要的列
        required_columns = ['门店编码', '门店名称', '城市']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            conn.close()
            return jsonify({'error': f'缺少必要列: {", ".join(missing_columns)}'}), 400
        
        success_count = 0
        error_count = 0
        error_details = []
        
        for index, row in df.iterrows():
            try:
                store_code = str(row['门店编码']).strip()
                store_name = str(row['门店名称']).strip()
                city = str(row['城市']).strip()
                address = str(row.get('地址', '')).strip() if pd.notna(row.get('地址')) else ''
                
                if not store_code or not store_name or not city:
                    error_count += 1
                    error_details.append(f'第{index+2}行: 门店编码、门店名称或城市不能为空')
                    continue
                
                # 检查门店编码是否已存在
                cursor.execute("SELECT id FROM stores WHERE store_code = ?", (store_code,))
                if cursor.fetchone():
                    # 更新现有门店
                    cursor.execute('''
                        UPDATE stores SET store_name = ?, store_city = ?, address = ?
                        WHERE store_code = ?
                    ''', (store_name, city, address, store_code))
                else:
                    # 插入新门店
                    cursor.execute('''
                        INSERT INTO stores (store_code, store_name, store_city, address)
                        VALUES (?, ?, ?, ?)
                    ''', (store_code, store_name, city, address))
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                error_details.append(f'第{index+2}行: {str(e)}')
        
        conn.commit()
        conn.close()
        
        result = {
            'success': True,
            'message': f'导入完成: 成功{success_count}条，失败{error_count}条',
            'success_count': success_count,
            'error_count': error_count
        }
        
        if error_details:
            result['error_details'] = error_details[:10]  # 只返回前10个错误
        
        return jsonify(result)
        
    except Exception as e:
        conn.close()
        return jsonify({'error': f'文件处理失败: {str(e)}'}), 500

@app.route('/api/timesheet', methods=['POST'])
@login_required
def add_timesheet():
    """添加工时记录"""
    data = request.get_json()
    
    required_fields = ['store_code', 'work_date', 'start_time', 'end_time']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'缺少必要字段: {field}'}), 400
    
    # 验证门店编码是否存在
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    cursor.execute("SELECT store_name FROM stores WHERE store_code = ?", (data['store_code'],))
    store = cursor.fetchone()
    if not store:
        conn.close()
        return jsonify({'error': '门店编码不存在'}), 400
    
    try:
        from datetime import datetime
        
        # 验证时间格式
        work_date = datetime.strptime(data['work_date'], '%Y-%m-%d').date()
        start_time = datetime.strptime(data['start_time'], '%H:%M').time()
        end_time = datetime.strptime(data['end_time'], '%H:%M').time()
        
        # 计算工作时长（小时）
        start_datetime = datetime.combine(work_date, start_time)
        end_datetime = datetime.combine(work_date, end_time)
        
        if end_datetime <= start_datetime:
            conn.close()
            return jsonify({'error': '结束时间必须晚于开始时间'}), 400
        
        work_hours = (end_datetime - start_datetime).total_seconds() / 3600
        
        # 检查是否已有相同日期和门店的记录
        cursor.execute('''
            SELECT id FROM timesheet_records 
            WHERE user_id = ? AND store_code = ? AND work_date = ?
        ''', (session['user_id'], data['store_code'], data['work_date']))
        
        existing = cursor.fetchone()
        if existing:
            conn.close()
            return jsonify({'error': '该日期该门店已有工时记录'}), 400
        
        # 插入工时记录
        cursor.execute('''
            INSERT INTO timesheet_records 
            (user_id, store_code, store_name, work_date, start_time, end_time, work_hours, work_content)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session['user_id'],
            data['store_code'],
            store[0],  # store_name
            data['work_date'],
            data['start_time'],
            data['end_time'],
            round(work_hours, 2),
            data.get('work_content', '')
        ))
        
        conn.commit()
        record_id = cursor.lastrowid
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '工时记录添加成功',
            'record_id': record_id,
            'work_hours': round(work_hours, 2)
        })
        
    except ValueError as e:
        conn.close()
        return jsonify({'error': f'时间格式错误: {str(e)}'}), 400
    except Exception as e:
        conn.close()
        return jsonify({'error': f'数据库错误: {str(e)}'}), 500

@app.route('/api/profile')
def profile():
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, username, email, name, department, position, created_at
        FROM users WHERE id = ?
    ''', (session['user_id'],))
    user = cursor.fetchone()
    conn.close()
    
    if user:
        return jsonify({
            'id': user[0],
            'username': user[1],
            'email': user[2],
            'name': user[3],
            'department': user[4],
            'position': user[5],
            'created_at': user[6]
        })
    else:
        return jsonify({'error': '用户信息不存在'}), 404

# API路由实现
@app.route('/api/users')
def api_users():
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, department, position FROM users ORDER BY department, name")
    users_data = cursor.fetchall()
    conn.close()
    
    return jsonify([{
        'id': row[0], 'name': row[1], 'department': row[2], 'position': row[3]
    } for row in users_data])

@app.route('/api/locations', methods=['GET', 'POST'])
def api_locations():
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    if request.method == 'GET':
        cursor.execute("SELECT id, name, address, longitude, latitude, geocoded FROM locations ORDER BY name")
        locations_data = cursor.fetchall()
        conn.close()
        
        return jsonify([{
            'id': row[0], 'name': row[1], 'address': row[2], 
            'longitude': row[3], 'latitude': row[4], 'geocoded': bool(row[5])
        } for row in locations_data])
    
    elif request.method == 'POST':
        data = request.json
        
        # 检查地点名称是否已存在
        cursor.execute("SELECT id FROM locations WHERE name = ?", (data['name'],))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': '地点名称已存在，请使用不同的名称'}), 400
        
        # 尝试获取坐标
        lng, lat = AmapService.geocode(data['address'])
        geocoded = lng is not None and lat is not None
        
        try:
            cursor.execute(
                "INSERT INTO locations (name, address, longitude, latitude, geocoded) VALUES (?, ?, ?, ?, ?)",
                (data['name'], data['address'], lng, lat, geocoded)
            )
            conn.commit()
            location_id = cursor.lastrowid
            conn.close()
            
            status_msg = "坐标获取成功" if geocoded else "坐标获取失败，将使用估算"
            
            return jsonify({
                'id': location_id,
                'name': data['name'],
                'address': data['address'],
                'longitude': lng,
                'latitude': lat,
                'geocoded': geocoded,
                'message': f'地点添加成功，{status_msg}'
            }), 201
        except Exception as e:
            conn.close()
            return jsonify({'error': f'数据库错误: {str(e)}'}), 500

@app.route('/api/stores', methods=['GET'])
def api_stores():
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    # 分页参数
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    search = request.args.get('search', '').strip()
    
    # 构建查询
    where_clause = ""
    params = []
    
    if search:
        where_clause = "WHERE store_code LIKE ? OR store_name LIKE ? OR store_city LIKE ?"
        search_pattern = f"%{search}%"
        params = [search_pattern, search_pattern, search_pattern]
    
    # 计算总数
    count_query = f"SELECT COUNT(*) FROM stores {where_clause}"
    cursor.execute(count_query, params)
    total = cursor.fetchone()[0]
    
    # 分页查询
    offset = (page - 1) * per_page
    query = f"""
        SELECT id, store_code, store_name, store_city, longitude, latitude, address, created_at 
        FROM stores {where_clause} 
        ORDER BY store_code 
        LIMIT ? OFFSET ?
    """
    cursor.execute(query, params + [per_page, offset])
    stores_data = cursor.fetchall()
    conn.close()
    
    # 返回分页数据
    return jsonify({
        'stores': [{
            'id': row[0], 'store_code': row[1], 'store_name': row[2], 'store_city': row[3],
            'longitude': row[4], 'latitude': row[5], 'address': row[6], 'created_at': row[7]
        } for row in stores_data],
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total,
            'pages': (total + per_page - 1) // per_page
        }
    })

@app.route('/api/stores/by-code/<store_code>', methods=['GET'])
def get_store_by_code(store_code):
    """根据门店编码获取门店信息"""
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, store_code, store_name, store_city, longitude, latitude, address
        FROM stores WHERE store_code = ?
    """, (store_code,))
    
    store_data = cursor.fetchone()
    conn.close()
    
    if store_data:
        return jsonify({
            'success': True,
            'store': {
                'id': store_data[0],
                'store_code': store_data[1], 
                'store_name': store_data[2],
                'store_city': store_data[3],
                'longitude': store_data[4],
                'latitude': store_data[5],
                'address': store_data[6]
            }
        })
    else:
        return jsonify({
            'success': False,
            'message': '门店编码不存在'
        }), 404

@app.route('/api/calculate-route', methods=['POST'])
def api_calculate_route():
    data = request.json
    
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    # 获取起点信息
    from_lng, from_lat, from_name = None, None, None
    if data.get('location_from_id'):
        cursor.execute("SELECT name, address, longitude, latitude FROM locations WHERE id = ?", 
                       (data['location_from_id'],))
        loc = cursor.fetchone()
        if loc:
            from_name, from_address, from_lng, from_lat = loc
    elif data.get('store_from_code'):
        cursor.execute("SELECT store_name, longitude, latitude FROM stores WHERE store_code = ?", 
                       (data['store_from_code'],))
        store = cursor.fetchone()
        if store:
            from_name, from_lng, from_lat = store
        else:
            return jsonify({'error': f'未找到门店编码: {data["store_from_code"]}'}), 400
    
    # 获取终点信息
    to_lng, to_lat, to_name = None, None, None
    if data.get('location_to_id'):
        cursor.execute("SELECT name, address, longitude, latitude FROM locations WHERE id = ?", 
                       (data['location_to_id'],))
        loc = cursor.fetchone()
        if loc:
            to_name, to_address, to_lng, to_lat = loc
    elif data.get('store_to_code'):
        cursor.execute("SELECT store_name, longitude, latitude FROM stores WHERE store_code = ?", 
                       (data['store_to_code'],))
        store = cursor.fetchone()
        if store:
            to_name, to_lng, to_lat = store
        else:
            return jsonify({'error': f'未找到门店编码: {data["store_to_code"]}'}), 400
    
    conn.close()
    
    if not (from_lng and from_lat and to_lng and to_lat):
        return jsonify({'error': '缺少坐标信息'}), 400
    
    transport_mode = data.get('transport_mode', '自驾')
    route_strategy = data.get('route_strategy', 'fastest')  # 默认时间最短
    
    # 计算距离和时间
    distance, travel_time, api_used = AmapService.calculate_distance_and_time(
        from_lng, from_lat, to_lng, to_lat, transport_mode, route_strategy
    )
    
    return jsonify({
        'distance': distance,
        'travel_time': travel_time,
        'from_name': from_name,
        'to_name': to_name,
        'transport_mode': transport_mode,
        'api_used': api_used
    })

@app.route('/api/timesheet', methods=['GET', 'POST'])
def api_timesheet():
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    if request.method == 'GET':
        user_id = request.args.get('user_id')
        
        cursor.execute('''
            SELECT tr.id, tr.date, 
                   lf.name as location_from_name, lt.name as location_to_name,
                   sf.store_name as store_from_name, st.store_name as store_to_name,
                   tr.transport_mode, tr.distance, tr.travel_time, 
                   tr.work_hours, tr.notes, tr.api_used
            FROM timesheet_records tr
            LEFT JOIN locations lf ON tr.location_from_id = lf.id
            LEFT JOIN locations lt ON tr.location_to_id = lt.id
            LEFT JOIN stores sf ON tr.store_from_id = sf.id
            LEFT JOIN stores st ON tr.store_to_id = st.id
            WHERE tr.user_id = ?
            ORDER BY tr.date DESC
            LIMIT 50
        ''', (user_id,))
        
        records = []
        for row in cursor.fetchall():
            from_name = row[2] or row[4]  # 地点名称或门店名称
            to_name = row[3] or row[5]    # 地点名称或门店名称
            
            records.append({
                'id': row[0], 'date': row[1],
                'from_name': from_name, 'to_name': to_name,
                'transport_mode': row[6], 'distance': row[7], 'travel_time': row[8],
                'work_hours': row[9], 'notes': row[10], 'api_used': bool(row[11])
            })
        
        conn.close()
        return jsonify(records)
    
    elif request.method == 'POST':
        data = request.json
        
        # 自动计算距离和时间
        distance, travel_time, api_used = None, None, False
        
        # 获取起点坐标
        from_lng, from_lat = None, None
        if data.get('location_from_id'):
            cursor.execute("SELECT longitude, latitude FROM locations WHERE id = ?", (data['location_from_id'],))
            coords = cursor.fetchone()
            if coords:
                from_lng, from_lat = coords
        elif data.get('store_from_code'):
            cursor.execute("SELECT longitude, latitude FROM stores WHERE store_code = ?", (data['store_from_code'],))
            coords = cursor.fetchone()
            if coords:
                from_lng, from_lat = coords
        
        # 获取终点坐标
        to_lng, to_lat = None, None
        if data.get('location_to_id'):
            cursor.execute("SELECT longitude, latitude FROM locations WHERE id = ?", (data['location_to_id'],))
            coords = cursor.fetchone()
            if coords:
                to_lng, to_lat = coords
        elif data.get('store_to_code'):
            cursor.execute("SELECT longitude, latitude FROM stores WHERE store_code = ?", (data['store_to_code'],))
            coords = cursor.fetchone()
            if coords:
                to_lng, to_lat = coords
        
        # 如果有坐标，计算距离和时间
        if from_lng and from_lat and to_lng and to_lat:
            distance, travel_time, api_used = AmapService.calculate_distance_and_time(
                from_lng, from_lat, to_lng, to_lat, data.get('transport_mode', '自驾')
            )
        
        # 查找门店ID（如果使用门店编码）
        store_from_id = None
        store_to_id = None
        
        if data.get('store_from_code'):
            cursor.execute("SELECT id FROM stores WHERE store_code = ?", (data['store_from_code'],))
            result = cursor.fetchone()
            if result:
                store_from_id = result[0]
                
        if data.get('store_to_code'):
            cursor.execute("SELECT id FROM stores WHERE store_code = ?", (data['store_to_code'],))
            result = cursor.fetchone()
            if result:
                store_to_id = result[0]
        
        # 插入记录
        try:
            cursor.execute('''
                INSERT INTO timesheet_records 
                (user_id, date, location_from_id, location_to_id, store_from_id, store_to_id,
                 transport_mode, distance, travel_time, work_hours, notes, api_used)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['user_id'], data['date'], 
                data.get('location_from_id'), data.get('location_to_id'),
                store_from_id, store_to_id,
                data.get('transport_mode', '自驾'),
                distance, travel_time, data['work_hours'], data.get('notes'), api_used
            ))
            
            conn.commit()
            record_id = cursor.lastrowid
            conn.close()
            
            return jsonify({
                'id': record_id,
                'message': '工时记录保存成功', 
                'distance': distance, 
                'travel_time': travel_time,
                'api_used': api_used
            }), 201
            
        except Exception as e:
            conn.close()
            return jsonify({'error': f'保存失败: {str(e)}'}), 500

@app.route('/api/statistics/<int:user_id>')
def api_statistics(user_id):
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    # 计算月度统计
    start_date = f"{month}-01"
    
    # 计算当月的最后一天
    year, month_num = map(int, month.split('-'))
    if month_num == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month_num + 1
    end_date = f"{next_year}-{next_month:02d}-01"
    
    cursor.execute('''
        SELECT 
            COUNT(*) as record_count,
            SUM(work_hours) as total_work_hours,
            SUM(travel_time) as total_travel_time,
            SUM(distance) as total_distance,
            COUNT(DISTINCT date) as work_days,
            COUNT(CASE WHEN api_used = 1 THEN 1 END) as api_records
        FROM timesheet_records 
        WHERE user_id = ? AND date >= ? AND date < ?
    ''', (user_id, start_date, end_date))
    
    stats = cursor.fetchone()
    conn.close()
    
    return jsonify({
        'user_id': user_id,
        'month': month,
        'total_work_hours': round(stats[1] or 0, 1),
        'total_travel_time': round(stats[2] or 0, 1),
        'total_distance': round(stats[3] or 0, 1),
        'work_days': stats[4] or 0,
        'avg_daily_hours': round((stats[1] or 0) / max(stats[4] or 1, 1), 1),
        'records_count': stats[0] or 0,
        'api_records': stats[5] or 0,
        'api_success_rate': round((stats[5] or 0) / max(stats[0] or 1, 1) * 100, 1)
    })

@app.route('/api/export/timesheet')
def export_timesheet():
    user_id = request.args.get('user_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    # 构建查询
    query = '''
        SELECT u.name, u.department, tr.date, 
               COALESCE(lf.name, sf.store_name) as from_name,
               COALESCE(lt.name, st.store_name) as to_name,
               tr.transport_mode, tr.distance, tr.travel_time, 
               tr.work_hours, tr.notes, tr.api_used
        FROM timesheet_records tr
        LEFT JOIN users u ON tr.user_id = u.id
        LEFT JOIN locations lf ON tr.location_from_id = lf.id
        LEFT JOIN locations lt ON tr.location_to_id = lt.id
        LEFT JOIN stores sf ON tr.store_from_id = sf.id
        LEFT JOIN stores st ON tr.store_to_id = st.id
        WHERE 1=1
    '''
    
    params = []
    if user_id:
        query += " AND tr.user_id = ?"
        params.append(user_id)
    if start_date:
        query += " AND tr.date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND tr.date <= ?"
        params.append(end_date)
    
    query += " ORDER BY u.name, tr.date"
    
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()
    
    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "工时记录"
    
    # 设置标题行
    headers = ['姓名', '部门', '日期', '出发地', '目的地', '交通方式', '距离(km)', '行程时间(h)', '工作时长(h)', '备注', '数据源']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 填充数据
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record[0])  # 姓名
        ws.cell(row=row, column=2, value=record[1])  # 部门
        ws.cell(row=row, column=3, value=record[2])  # 日期
        ws.cell(row=row, column=4, value=record[3] or '-')  # 出发地
        ws.cell(row=row, column=5, value=record[4] or '-')  # 目的地
        ws.cell(row=row, column=6, value=record[5])  # 交通方式
        ws.cell(row=row, column=7, value=record[6] or 0)  # 距离
        ws.cell(row=row, column=8, value=record[7] or 0)  # 行程时间
        ws.cell(row=row, column=9, value=record[8])  # 工作时长
        ws.cell(row=row, column=10, value=record[9] or '-')  # 备注
        ws.cell(row=row, column=11, value='高德API' if record[10] else '智能估算')  # 数据源
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回文件
    return send_file(
        output,
        as_attachment=True,
        download_name=f'工时表_{start_date}_{end_date}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/export/stores')
def export_stores():
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    cursor.execute("SELECT store_code, store_name, store_city, longitude, latitude, address FROM stores ORDER BY store_code")
    stores = cursor.fetchall()
    conn.close()
    
    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "门店信息"
    
    # 设置标题行
    headers = ['门店编码', '门店名称', '所在城市', '经度', '纬度', '详细地址']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 填充数据
    for row, store in enumerate(stores, 2):
        for col, value in enumerate(store, 1):
            ws.cell(row=row, column=col, value=value)
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name='门店信息.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/export/json')
def export_json():
    user_id = request.args.get('user_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect('enhanced_timesheet.db')
    cursor = conn.cursor()
    
    # 导出工时记录
    query = '''
        SELECT u.name, u.department, tr.date, 
               COALESCE(lf.name, sf.store_name) as from_name,
               COALESCE(lt.name, st.store_name) as to_name,
               tr.transport_mode, tr.distance, tr.travel_time, 
               tr.work_hours, tr.notes, tr.api_used
        FROM timesheet_records tr
        LEFT JOIN users u ON tr.user_id = u.id
        LEFT JOIN locations lf ON tr.location_from_id = lf.id
        LEFT JOIN locations lt ON tr.location_to_id = lt.id
        LEFT JOIN stores sf ON tr.store_from_id = sf.id
        LEFT JOIN stores st ON tr.store_to_id = st.id
        WHERE 1=1
    '''
    
    params = []
    if user_id:
        query += " AND tr.user_id = ?"
        params.append(user_id)
    if start_date:
        query += " AND tr.date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND tr.date <= ?"
        params.append(end_date)
    
    query += " ORDER BY u.name, tr.date"
    
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()
    
    # 转换为JSON格式
    json_data = []
    for record in records:
        json_data.append({
            'name': record[0],
            'department': record[1],
            'date': record[2],
            'from_name': record[3],
            'to_name': record[4],
            'transport_mode': record[5],
            'distance': record[6],
            'travel_time': record[7],
            'work_hours': record[8],
            'notes': record[9],
            'api_used': bool(record[10])
        })
    
    # 创建JSON文件
    output = io.BytesIO()
    output.write(json.dumps(json_data, ensure_ascii=False, indent=2).encode('utf-8'))
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'工时数据_{start_date}_{end_date}.json',
        mimetype='application/json'
    )

if __name__ == '__main__':
    print("🚀 智能工时表管理系统 - 增强版启动中...")
    print(f"🗺️ 高德地图API: {AMAP_API_KEY[:8]}...")
    print("🏪 门店管理功能已集成")
    print("📊 Excel导出功能已就绪")
    print("🔧 具备智能降级机制，确保系统稳定运行")
    
    init_db()
    
    print("✅ 系统就绪！")
    print("🌐 访问地址: http://localhost:8080")
    print("📱 支持桌面和移动设备")
    print("📋 功能包括：工时管理 + 门店管理 + 数据导出")
    print("-" * 50)
    
    app.run(debug=True, host='0.0.0.0', port=8080)
